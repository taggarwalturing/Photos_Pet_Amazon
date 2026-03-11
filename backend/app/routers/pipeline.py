"""
Master Pipeline Control API
============================
Admin endpoints for controlling and monitoring the master pipeline.
"""

import asyncio
import json
import os
from pathlib import Path
from datetime import datetime
from typing import Optional, List, Dict
from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks, UploadFile, File, Query
from sqlalchemy.orm import Session
from sqlalchemy import text, func
from pydantic import BaseModel

from app.database import get_db
from app.dependencies import require_admin
from app.models.user import User
from app.models.image import Image
from app.models.drive_folder import DriveFolder


def _get_pipeline_workspace() -> Path:
    """Return the pipeline workspace path from env, falling back to the legacy local path."""
    from app.utils import get_pipeline_workspace
    return get_pipeline_workspace()

router = APIRouter(prefix="/admin/pipeline", tags=["Master Pipeline"])

# Pipeline status storage (in-memory for now)
pipeline_status = {
    "is_running": False,
    "current_step": None,
    "progress": {
        "download": {"status": "pending", "current": 0, "total": 0, "message": ""},
        "deduplicate": {"status": "pending", "current": 0, "total": 0, "message": ""},
        "biometric": {"status": "pending", "current": 0, "total": 0, "message": ""}
    },
    "started_at": None,
    "completed_at": None,
    "errors": [],
    "summary": {},
    # Per-folder tracking
    "current_folder": None,
    "current_folder_idx": 0,
    "total_folders": 0,
    "folder_progress": {}  # { folder_id: { status, steps: {download, deduplicate, biometric} } }
}


class PipelineRunRequest(BaseModel):
    download: bool = False
    deduplicate: bool = False
    biometric: bool = True
    use_llm: bool = False
    threshold: float = 0.85
    folder_ids: Optional[List[str]] = None  # Specific folder IDs to process (uses DB folders if None)
    source: str = "gcs"  # Image source: "gcs" (default) or "drive"
    force_reprocess: bool = False  # If True, re-run ALL completed folders
    force_reprocess_folder_ids: Optional[List[str]] = None  # Specific completed folders to force-reprocess


class ReprocessRequest(BaseModel):
    image_ids: List[int]


@router.get("/status")
def get_pipeline_status(
    admin: User = Depends(require_admin)
):
    """Get current pipeline execution status."""
    return pipeline_status


@router.post("/start")
async def start_pipeline(
    request: PipelineRunRequest,
    background_tasks: BackgroundTasks,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Start the master pipeline with specified steps."""
    global pipeline_status
    
    if pipeline_status["is_running"]:
        raise HTTPException(status_code=400, detail="Pipeline is already running")
    
    # Determine folder IDs: explicit request > DB folders > env config
    folder_ids = request.folder_ids
    if not folder_ids and request.download:
        if request.source == "gcs":
            # For GCS: discover folders from bucket prefixes
            folder_ids = _resolve_gcs_folders(db)
        else:
            # Pull folder IDs from database
            db_folders = db.query(DriveFolder).filter(DriveFolder.status != "disabled").all()
            if db_folders:
                folder_ids = [f.folder_id for f in db_folders]

    # Resolve parent folders into child subfolders BEFORE launching the pipeline
    # so the initial status response already shows the correct leaf folders.
    if folder_ids and request.download and request.source == "drive":
        folder_ids = _resolve_parent_folders(folder_ids, db)

    # ── Skip already-completed folders (unless force-reprocessed or new images detected) ──
    # force_reprocess=True  → force ALL completed folders
    # force_reprocess_folder_ids=[…] → force only those specific folders
    # Smart detection: if GCS has more images than DB for a completed folder, re-include it
    skipped_folders = []
    folders_with_new_images = []
    force_ids = set(request.force_reprocess_folder_ids or [])
    if folder_ids and not request.force_reprocess:
        completed_ids = set()
        for rec in db.query(DriveFolder).filter(
            DriveFolder.folder_id.in_(folder_ids),
            DriveFolder.status == "completed",
        ).all():
            # Skip this folder only if it's NOT in the per-folder force list
            if rec.folder_id not in force_ids:
                completed_ids.add(rec.folder_id)

        if completed_ids and request.source == "gcs" and request.download:
            # Smart detection: check if GCS has new images not yet in DB
            try:
                from app.utils.gcs import list_blobs
                image_exts = {'.jpg', '.jpeg', '.png', '.gif', '.webp', '.heic', '.heif', '.avif', '.bmp', '.tiff', '.tif'}
                for fid in list(completed_ids):
                    # Count images in GCS input/{fid}/
                    gcs_prefix = f"input/{fid}/"
                    gcs_blobs = [
                        b for b in list_blobs(gcs_prefix)
                        if os.path.splitext(b)[1].lower() in image_exts
                    ]
                    gcs_count = len(gcs_blobs)

                    # Count images in DB for this folder
                    db_count = db.query(func.count(Image.id)).filter(
                        Image.source_folder_id == fid,
                    ).scalar() or 0

                    if gcs_count > db_count:
                        # New images detected — don't skip this folder
                        completed_ids.discard(fid)
                        folders_with_new_images.append(fid)
                        print(f"[PIPELINE] Folder {fid}: {gcs_count} in GCS vs {db_count} in DB — "
                              f"new images detected, will re-process")
            except Exception as e:
                print(f"[PIPELINE] Smart detection failed, falling back to skip: {e}")

        if completed_ids:
            skipped_folders = sorted(completed_ids)
            folder_ids = [fid for fid in folder_ids if fid not in completed_ids]
            print(f"[PIPELINE] Skipping {len(skipped_folders)} already-completed folder(s): {skipped_folders}")
        if folders_with_new_images:
            print(f"[PIPELINE] Re-including {len(folders_with_new_images)} folder(s) with new images: {folders_with_new_images}")

    if not folder_ids:
        msg = "No new folders to process."
        if skipped_folders:
            msg += f" {len(skipped_folders)} folder(s) already completed — enable 'Force Reprocess' to re-run them."
        raise HTTPException(status_code=400, detail=msg)

    # Reset status
    pipeline_status = {
        "is_running": True,
        "current_step": None,
        "progress": {
            "download": {"status": "pending", "current": 0, "total": 0, "message": ""},
            "deduplicate": {"status": "pending", "current": 0, "total": 0, "message": ""},
            "biometric": {"status": "pending", "current": 0, "total": 0, "message": ""}
        },
        "started_at": datetime.now().isoformat(),
        "completed_at": None,
        "errors": [],
        "summary": {},
        "requested_by": admin.username,
        "current_folder": None,
        "current_folder_idx": 0,
        "total_folders": len(folder_ids),
        "folder_progress": {},
        "skipped_folders": skipped_folders,
    }
    
    # Pre-populate folder_progress so frontend can show them immediately
    if folder_ids:
        for fid in folder_ids:
            pipeline_status["folder_progress"][fid] = {
                "status": "pending",
                "current_step": None,
                "steps": {
                    "download": {"status": "pending", "current": 0, "total": 0, "message": ""},
                    "deduplicate": {"status": "pending", "current": 0, "total": 0, "message": ""},
                    "biometric": {"status": "pending", "current": 0, "total": 0, "message": ""},
                },
                "errors": []
            }
    
    # Run pipeline in background
    background_tasks.add_task(
        run_pipeline_background,
        request.download,
        request.deduplicate,
        request.biometric,
        request.use_llm,
        request.threshold,
        db,
        folder_ids,
        request.source,
    )
    
    return {"message": "Pipeline started successfully", "status": pipeline_status}


@router.post("/stop")
def stop_pipeline(
    admin: User = Depends(require_admin)
):
    """Stop the currently running pipeline."""
    global pipeline_status
    
    if not pipeline_status["is_running"]:
        raise HTTPException(status_code=400, detail="No pipeline is currently running")
    
    # TODO: Implement graceful pipeline termination
    pipeline_status["is_running"] = False
    pipeline_status["current_step"] = "stopped"
    pipeline_status["completed_at"] = datetime.now().isoformat()
    
    return {"message": "Pipeline stop requested", "status": pipeline_status}


@router.get("/errors")
def get_pipeline_errors(
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Get images that failed during processing."""
    
    # Get images with processing errors
    failed_images = db.execute(text("""
        SELECT id, filename, compliance_status, pipeline_status, human_faces_detected
        FROM images
        WHERE compliance_status IN ('failed', 'error', 'needs_reprocess')
        OR pipeline_status IN ('failed', 'error')
        ORDER BY id DESC
    """)).fetchall()
    
    return {
        "total_errors": len(failed_images),
        "errors": [
            {
                "image_id": img[0],
                "filename": img[1],
                "status": img[2],
                "log": img[3] or "",
                "faces_detected": img[4]
            }
            for img in failed_images
        ]
    }


@router.post("/reprocess")
async def reprocess_failed_images(
    request: ReprocessRequest,
    background_tasks: BackgroundTasks,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Reprocess specific images that failed."""
    
    if pipeline_status["is_running"]:
        raise HTTPException(
            status_code=400,
            detail="Cannot reprocess while pipeline is running"
        )
    
    # Verify images exist
    images = db.query(Image).filter(Image.id.in_(request.image_ids)).all()
    
    if len(images) != len(request.image_ids):
        raise HTTPException(
            status_code=404,
            detail=f"Some images not found. Found {len(images)} of {len(request.image_ids)}"
        )
    
    # Reset their processing status
    for image in images:
        image.pipeline_status = "pending"
        image.compliance_status = "pending_reprocess"
    
    db.commit()
    
    # Start reprocessing in background
    background_tasks.add_task(
        reprocess_images_background,
        request.image_ids,
        db
    )
    
    return {
        "message": f"Reprocessing {len(images)} images",
        "image_ids": request.image_ids
    }


class ReprocessFolderRequest(BaseModel):
    folder_id: str
    reprocess_all: bool = False  # If True, reprocess ALL images; if False, only failed/pending


@router.post("/reprocess-folder")
async def reprocess_folder(
    request: ReprocessFolderRequest,
    background_tasks: BackgroundTasks,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """
    Reprocess images in a specific folder.
    By default only reprocesses failed/pending images.
    Set reprocess_all=True to reprocess everything in the folder.
    """
    if pipeline_status["is_running"]:
        raise HTTPException(status_code=400, detail="Cannot reprocess while pipeline is running")

    if request.reprocess_all:
        images = db.query(Image).filter(
            Image.source_folder_id == request.folder_id,
            Image.is_duplicate == False,  # noqa: E712
        ).all()
    else:
        images = db.query(Image).filter(
            Image.source_folder_id == request.folder_id,
            Image.is_duplicate == False,  # noqa: E712
            Image.pipeline_status.in_(["pending", "failed", "error", None]),
        ).all()

    if not images:
        raise HTTPException(
            status_code=404,
            detail=f"No {'images' if request.reprocess_all else 'failed/pending images'} found in folder {request.folder_id}"
        )

    image_ids = [img.id for img in images]

    # Reset their processing status
    for image in images:
        image.pipeline_status = "pending"
        image.compliance_status = "pending_reprocess"
    db.commit()

    # Start reprocessing in background
    background_tasks.add_task(reprocess_images_background, image_ids, db)

    return {
        "message": f"Reprocessing {len(images)} images in folder {request.folder_id}",
        "image_ids": image_ids,
        "folder_id": request.folder_id,
    }


@router.get("/summary")
def get_pipeline_summary(
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Get overall pipeline statistics."""
    
    stats = db.execute(text("""
        SELECT 
            COUNT(*) as total,
            SUM(CASE WHEN pipeline_status = 'completed' THEN 1 ELSE 0 END) as processed,
            SUM(CASE WHEN compliance_status = 'clean' THEN 1 ELSE 0 END) as clean,
            SUM(CASE WHEN compliance_status IN ('processed', 'blurred', 'obfuscated') THEN 1 ELSE 0 END) as blurred,
            SUM(CASE WHEN compliance_status IN ('failed', 'error') THEN 1 ELSE 0 END) as failed,
            SUM(CASE WHEN human_faces_detected > 0 THEN 1 ELSE 0 END) as with_faces
        FROM images
    """)).fetchone()
    
    return {
        "total_images": stats[0] or 0,
        "processed": stats[1] or 0,
        "clean": stats[2] or 0,
        "blurred": stats[3] or 0,
        "failed": stats[4] or 0,
        "with_faces": stats[5] or 0,
        "pending": (stats[0] or 0) - (stats[1] or 0)
    }


def _load_all_drive_metadata():
    """
    Load and aggregate drive_metadata.json from all per-folder workspaces.
    Falls back to legacy root workspace metadata if no per-folder workspaces exist.
    """
    workspace = _get_pipeline_workspace()
    folders_dir = workspace / "folders"
    
    aggregated = {
        "total_in_drive": 0,
        "unique_filenames": 0,
        "duplicate_filename_count": 0,
        "duplicate_filenames": {},
        "scanned_at": "",
        "per_folder": {},
    }
    
    per_folder_found = False
    
    if folders_dir.exists():
        for folder_dir in sorted(folders_dir.iterdir()):
            if not folder_dir.is_dir():
                continue
            # Prefer GCS metadata, fall back to Drive metadata
            meta_path = folder_dir / "gcs_metadata.json"
            if not meta_path.exists():
                meta_path = folder_dir / "drive_metadata.json"
            if meta_path.exists():
                try:
                    with open(meta_path, 'r') as f:
                        meta = json.load(f)
                    fid = meta.get("folder_id", folder_dir.name)
                    total = meta.get("total_in_gcs") or meta.get("total_in_drive", 0)
                    unique = meta.get("unique_filenames", 0)
                    dup_count = meta.get("duplicate_filename_count", 0)
                    aggregated["total_in_drive"] += total
                    aggregated["unique_filenames"] += unique
                    aggregated["duplicate_filename_count"] += dup_count
                    for dn, dc in meta.get("duplicate_filenames", {}).items():
                        aggregated["duplicate_filenames"][dn] = aggregated["duplicate_filenames"].get(dn, 0) + dc
                    if meta.get("scanned_at", "") > aggregated["scanned_at"]:
                        aggregated["scanned_at"] = meta["scanned_at"]
                    aggregated["per_folder"][fid] = {
                        "total": total,
                        "unique": unique,
                        "duplicates": dup_count,
                    }
                    per_folder_found = True
                except Exception:
                    pass
    
    # Fallback: legacy root workspace
    if not per_folder_found:
        legacy_path = workspace / "drive_metadata.json"
        if legacy_path.exists():
            try:
                with open(legacy_path, 'r') as f:
                    legacy = json.load(f)
                aggregated["total_in_drive"] = legacy.get("total_in_drive", 0)
                aggregated["unique_filenames"] = legacy.get("unique_filenames", 0)
                aggregated["duplicate_filename_count"] = legacy.get("duplicate_filename_count", 0)
                aggregated["duplicate_filenames"] = legacy.get("duplicate_filenames", {})
                aggregated["scanned_at"] = legacy.get("scanned_at", "")
                aggregated["per_folder"] = legacy.get("per_folder", {})
            except Exception:
                pass
    
    return aggregated


def _get_workspace_dedup_stats():
    """Get dedup stats by scanning per-folder workspaces (or legacy workspace)."""
    workspace = _get_pipeline_workspace()
    folders_dir = workspace / "folders"
    extensions = {'.jpg', '.jpeg', '.png', '.webp', '.heic', '.heif', '.avif'}
    
    unique_count = 0
    duplicate_count = 0
    cluster_count = 0
    
    try:
        # Read deduplication stats from JSON files (DB-driven approach)
        if folders_dir.exists():
            for folder_dir in folders_dir.iterdir():
                if not folder_dir.is_dir():
                    continue
                dedup_stats_path = folder_dir / "deduplication_stats.json"
                if dedup_stats_path.exists():
                    import json as _json
                    with open(dedup_stats_path, 'r') as f:
                        dedup_stats = _json.load(f)
                    unique_count += dedup_stats.get('unique_images', 0)
                    duplicate_count += dedup_stats.get('duplicate_images', 0)
                else:
                    # Fallback: count deliverable/ images as unique
                    d_dir = folder_dir / "deliverable"
                    if d_dir.exists():
                        unique_count += len([f for f in d_dir.iterdir() if f.is_file() and f.suffix.lower() in extensions])
        
        # Fallback: legacy workspace
        if unique_count == 0 and duplicate_count == 0:
            dedup_stats_path = workspace / "deduplication_stats.json"
            if dedup_stats_path.exists():
                import json as _json
                with open(dedup_stats_path, 'r') as f:
                    dedup_stats = _json.load(f)
                unique_count = dedup_stats.get('unique_images', 0)
                duplicate_count = dedup_stats.get('duplicate_images', 0)
    except Exception as e:
        print(f"Error reading deduplication stats: {e}")
    
    return unique_count, duplicate_count, cluster_count


@router.get("/stats")
def get_pipeline_stats(
    db: Session = Depends(get_db)
):
    """Get detailed pipeline statistics for UI display."""
    
    # Get image stats from database
    stats = db.execute(text("""
        SELECT 
            COUNT(*) as total,
            SUM(CASE WHEN pipeline_status = 'completed' THEN 1 ELSE 0 END) as processed,
            SUM(CASE WHEN pipeline_status != 'completed' OR pipeline_status IS NULL THEN 1 ELSE 0 END) as pending,
            SUM(CASE WHEN compliance_status IN ('failed', 'error') THEN 1 ELSE 0 END) as failed,
            SUM(CASE WHEN human_faces_detected > 0 OR compliance_status IN ('blurred', 'processed', 'obfuscated') THEN 1 ELSE 0 END) as with_faces,
            SUM(CASE WHEN (human_faces_detected = 0 OR human_faces_detected IS NULL) AND compliance_status = 'clean' THEN 1 ELSE 0 END) as without_faces,
            SUM(CASE WHEN compliance_status LIKE '%screenshot%' THEN 1 ELSE 0 END) as screenshots
        FROM images
    """)).fetchone()
    
    total = stats[0] or 0
    processed = stats[1] or 0
    pending = stats[2] or 0
    failed = stats[3] or 0

    # Dedup stats from DB (images table)
    duplicate_count = db.execute(
        text("SELECT COUNT(*) FROM images WHERE is_duplicate = TRUE")
    ).scalar() or 0
    unique_count = total - duplicate_count

    # Aggregate drive_folders stats from DB
    folder_agg = db.execute(text("""
        SELECT
            COALESCE(SUM(total_in_drive), 0),
            COALESCE(SUM(unique_count), 0),
            COALESCE(SUM(duplicate_count), 0),
            MAX(last_run_at)
        FROM drive_folders
    """)).fetchone()

    total_in_gcs = folder_agg[0] or 0
    gcs_unique = folder_agg[1] or 0
    gcs_duplicates = folder_agg[2] or 0
    last_run_at = folder_agg[3]

    return {
        "total_images": total,
        "processed": processed,
        "pending": pending,
        "failed": failed,
        "unique_images": unique_count,
        "duplicate_images": duplicate_count,
        "duplicate_clusters": 0,
        "images_with_faces": stats[4] or 0,
        "images_without_faces": stats[5] or 0,
        "screenshots_skipped": stats[6] or 0,
        "status": "idle" if not pipeline_status["is_running"] else pipeline_status["current_step"],
        "last_run": pipeline_status.get("completed_at") or (last_run_at.isoformat() if last_run_at else None),
        # GCS / Drive metadata (aggregated from drive_folders table)
        "total_in_drive": total_in_gcs,
        "drive_unique_filenames": gcs_unique if gcs_unique > 0 else unique_count,
        "drive_duplicate_filenames": gcs_duplicates,
        "drive_duplicate_details": {},
        "drive_scanned_at": last_run_at.isoformat() if last_run_at else "",
    }


# Background task functions


def _update_drive_folder_stats(folder_ids: List[str] = None):
    """Update DriveFolder rows in DB with actual stats from pipeline output and images table."""
    from app.database import SessionLocal
    from pathlib import Path
    import json

    if not folder_ids:
        return

    db = SessionLocal()
    try:
        pipeline_workspace = _get_pipeline_workspace()

        for fid in folder_ids:
            folder_record = db.query(DriveFolder).filter(DriveFolder.folder_id == fid).first()
            if not folder_record:
                continue

            ws = pipeline_workspace / "folders" / fid

            # Drive or GCS metadata — try local file first, then query GCS directly
            meta_path = ws / "gcs_metadata.json"
            if not meta_path.exists():
                meta_path = ws / "drive_metadata.json"
            if meta_path.exists():
                try:
                    with open(meta_path) as f:
                        dm = json.load(f)
                    folder_record.total_in_drive = dm.get("total_in_gcs") or dm.get("total_in_drive", 0)
                except Exception:
                    pass
            elif (folder_record.total_in_drive or 0) == 0:
                # Fallback: count blobs in GCS directly
                try:
                    from app.utils.gcs import list_blobs
                    blobs = [b for b in list_blobs(f"input/{fid}/") if not b.endswith("/")]
                    folder_record.total_in_drive = len(blobs)
                except Exception:
                    pass

            # Count images in DB for this folder
            from sqlalchemy import text
            row = db.execute(
                text("SELECT COUNT(*) FROM images WHERE source_folder_id = :fid"),
                {"fid": fid}
            ).scalar()
            folder_record.downloaded_count = row or 0

            # Count blurred / clean from DB
            blurred = db.execute(
                text("SELECT COUNT(*) FROM images WHERE source_folder_id = :fid AND compliance_status = 'blurred'"),
                {"fid": fid}
            ).scalar()
            clean = db.execute(
                text("SELECT COUNT(*) FROM images WHERE source_folder_id = :fid AND compliance_status = 'clean'"),
                {"fid": fid}
            ).scalar()
            folder_record.blurred_count = blurred or 0
            folder_record.clean_count = clean or 0

            # Dedup stats — try local file first, fallback to DB counts
            dedup_path = ws / "deduplication_stats.json"
            if dedup_path.exists():
                try:
                    with open(dedup_path) as f:
                        dd = json.load(f)
                    folder_record.unique_count = dd.get("unique_images", 0)
                    folder_record.duplicate_count = dd.get("duplicates_found", 0)
                except Exception:
                    pass
            # If still not set, derive from DB counts
            if (folder_record.unique_count or 0) == 0 and (folder_record.downloaded_count or 0) > 0:
                dup_in_db = db.execute(
                    text("SELECT COUNT(*) FROM images WHERE source_folder_id = :fid AND is_duplicate = TRUE"),
                    {"fid": fid}
                ).scalar() or 0
                folder_record.unique_count = (folder_record.downloaded_count or 0) - dup_in_db
                folder_record.duplicate_count = dup_in_db

            # Obfuscation stats
            obf_path = ws / "obfuscation_results.json"
            if obf_path.exists():
                try:
                    with open(obf_path) as f:
                        od = json.load(f)
                    stats = od.get("statistics", {})
                    folder_record.failed_count = stats.get("verification_failed", 0) + stats.get("failed", 0)
                except Exception:
                    pass

            # Determine per-folder status from in-memory progress
            fp = pipeline_status.get("folder_progress", {}).get(fid, {})
            fp_status = fp.get("status", "completed")
            fp_errors = fp.get("errors", [])

            if fp_status == "failed" or fp_errors:
                folder_record.status = "failed"
                folder_record.error_log = "; ".join(fp_errors) if fp_errors else "Unknown error"
            else:
                folder_record.status = "completed"
                if (folder_record.downloaded_count or 0) == 0 and (folder_record.total_in_drive or 0) > 0:
                    folder_record.error_log = (
                        "All images already exist in DB from another folder"
                    )
                else:
                    folder_record.error_log = None

            folder_record.last_run_at = datetime.now()

        db.commit()
    except Exception as e:
        db.rollback()
        raise e
    finally:
        db.close()


def _init_folder_progress():
    """Create a fresh per-folder progress dict."""
    return {
        "status": "pending",
        "current_step": None,
        "steps": {
            "download":    {"status": "pending", "current": 0, "total": 0, "message": ""},
            "deduplicate": {"status": "pending", "current": 0, "total": 0, "message": ""},
            "biometric":   {"status": "pending", "current": 0, "total": 0, "message": ""},
        },
        "errors": [],
    }


def _resolve_gcs_folders(db: Session) -> List[str]:
    """
    Discover folder IDs by listing GCS bucket prefixes under ``input/``.

    Each sub-prefix (e.g. ``input/folder_A/``) is treated as a folder.
    New folders are registered in the ``drive_folders`` DB table automatically.

    Returns:
        List of folder IDs found in the GCS bucket.
    """
    try:
        from app.utils.gcs import list_prefixes
    except ImportError:
        print("[GCS_RESOLVE] google-cloud-storage not available — skipping GCS folder discovery")
        # Fallback to DB folders
        db_folders = db.query(DriveFolder).filter(DriveFolder.status != "disabled").all()
        return [f.folder_id for f in db_folders]

    try:
        prefixes = list_prefixes("input/")
        folder_ids = []
        for p in sorted(prefixes):
            # p looks like "input/folder_A/"  →  extract "folder_A"
            parts = p.strip("/").split("/")
            if len(parts) >= 2:
                fid = parts[1]
                folder_ids.append(fid)

                # Register in DB if not already present
                existing = db.query(DriveFolder).filter(DriveFolder.folder_id == fid).first()
                if not existing:
                    db.add(DriveFolder(
                        folder_id=fid,
                        folder_name=fid,
                        notes="Auto-discovered from GCS bucket",
                        status="pending",
                    ))

        if folder_ids:
            db.commit()
            print(f"[GCS_RESOLVE] Found {len(folder_ids)} folder(s) in GCS bucket: {folder_ids}")
        else:
            print("[GCS_RESOLVE] No folders found under input/ in GCS bucket")
            # Fallback to DB folders
            db_folders = db.query(DriveFolder).filter(DriveFolder.status != "disabled").all()
            folder_ids = [f.folder_id for f in db_folders]

        return folder_ids

    except Exception as e:
        print(f"[GCS_RESOLVE] Error listing GCS prefixes: {e}")
        # Fallback to DB folders
        db_folders = db.query(DriveFolder).filter(DriveFolder.status != "disabled").all()
        return [f.folder_id for f in db_folders]


def _resolve_parent_folders(folder_ids: List[str], db: Session) -> List[str]:
    """
    For each folder ID, check if it's a "parent" folder (contains only subfolders,
    no direct image files).  If so, expand it into its child subfolder IDs and
    register them in the drive_folders DB table.

    Returns the final flat list of leaf-folder IDs to process.
    """
    import os
    from dotenv import dotenv_values

    try:
        from google.oauth2 import service_account
        from googleapiclient.discovery import build
    except ImportError:
        print("[RESOLVE] Google Drive libraries not available — skipping parent-folder expansion")
        return folder_ids

    # Build Drive service (same logic as master_pipeline._get_drive_service)
    env_vals = dotenv_values(Path(__file__).parent.parent.parent / ".env")

    sa_file = env_vals.get("GOOGLE_SERVICE_ACCOUNT_FILE") or os.getenv("GOOGLE_SERVICE_ACCOUNT_FILE")
    if not sa_file:
        print("[RESOLVE] No GOOGLE_SERVICE_ACCOUNT_FILE — skipping parent-folder expansion")
        return folder_ids

    sa_path = Path(__file__).parent.parent.parent / sa_file
    if not sa_path.exists():
        sa_path = Path(sa_file)
    if not sa_path.exists():
        print(f"[RESOLVE] Service account file not found: {sa_file} — skipping")
        return folder_ids

    import json as _json
    with open(sa_path) as f:
        creds_dict = _json.load(f)

    creds = service_account.Credentials.from_service_account_info(
        creds_dict, scopes=["https://www.googleapis.com/auth/drive.readonly"]
    )
    service = build("drive", "v3", credentials=creds)

    IMAGE_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.gif', '.webp', '.heic', '.heif', '.avif', '.bmp', '.tiff', '.tif'}
    resolved: List[str] = []

    for fid in folder_ids:
        print(f"[RESOLVE] Checking folder: {fid}")
        try:
            # List immediate children
            query = f"'{fid}' in parents and trashed=false"
            all_children = []
            page_token = None
            while True:
                resp = service.files().list(
                    q=query, spaces="drive",
                    fields="nextPageToken, files(id, name, mimeType)",
                    pageToken=page_token, pageSize=200,
                    includeItemsFromAllDrives=True, supportsAllDrives=True,
                ).execute()
                all_children.extend(resp.get("files", []))
                page_token = resp.get("nextPageToken")
                if not page_token:
                    break

            subfolders = [c for c in all_children if c["mimeType"] == "application/vnd.google-apps.folder"]
            image_files = [c for c in all_children
                           if c["mimeType"] != "application/vnd.google-apps.folder"
                           and os.path.splitext(c["name"])[1].lower() in IMAGE_EXTENSIONS]

            if subfolders and not image_files:
                # This is a parent folder — expand into subfolders
                print(f"[RESOLVE] '{fid}' is a parent folder with {len(subfolders)} subfolder(s) — expanding")

                # Get parent folder name for notes
                try:
                    parent_meta = service.files().get(
                        fileId=fid, fields="name", supportsAllDrives=True
                    ).execute()
                    parent_name = parent_meta.get("name", fid)
                except Exception:
                    parent_name = fid

                for sf in sorted(subfolders, key=lambda x: x["name"]):
                    sf_id = sf["id"]
                    sf_name = sf["name"]
                    resolved.append(sf_id)
                    print(f"[RESOLVE]   → {sf_name} ({sf_id})")

                    # Register in DB if not already present
                    existing = db.query(DriveFolder).filter(DriveFolder.folder_id == sf_id).first()
                    if not existing:
                        db.add(DriveFolder(
                            folder_id=sf_id,
                            folder_name=sf_name,
                            notes=f"Auto-discovered from parent: {parent_name}",
                            status="pending",
                        ))

                # Remove the parent entry from DB if it was registered
                parent_entry = db.query(DriveFolder).filter(DriveFolder.folder_id == fid).first()
                if parent_entry:
                    db.delete(parent_entry)

                db.commit()
            else:
                # This folder directly contains images — keep as-is
                resolved.append(fid)
                if image_files:
                    print(f"[RESOLVE] '{fid}' contains {len(image_files)} image(s) — processing directly")
                else:
                    print(f"[RESOLVE] '{fid}' is empty or inaccessible — keeping as-is")

        except Exception as e:
            print(f"[RESOLVE] Error checking folder {fid}: {e} — keeping as-is")
            resolved.append(fid)

    # De-duplicate while preserving order
    seen = set()
    unique = []
    for r in resolved:
        if r not in seen:
            seen.add(r)
            unique.append(r)

    if len(unique) != len(folder_ids):
        print(f"[RESOLVE] Expanded {len(folder_ids)} input folder(s) → {len(unique)} leaf folder(s)")

    return unique


def run_pipeline_background(
    download: bool,
    deduplicate: bool,
    biometric: bool,
    use_llm: bool,
    threshold: float,
    db: Session,
    folder_ids: List[str] = None,
    source: str = "gcs",
):
    """Run the master pipeline in the background."""
    import subprocess, sys, re
    from pathlib import Path
    global pipeline_status

    try:
        print(f"[PIPELINE] Starting pipeline: download={download}, deduplicate={deduplicate}, biometric={biometric}, source={source}")

        # ── Resolve parent folders into child subfolders (Drive only) ──
        if folder_ids and download and source == "drive":
            original_count = len(folder_ids)
            folder_ids = _resolve_parent_folders(folder_ids, db)
            if len(folder_ids) != original_count:
                print(f"[PIPELINE] Expanded {original_count} → {len(folder_ids)} folder(s) after parent discovery")
                # Re-populate folder_progress for the expanded list
                pipeline_status["total_folders"] = len(folder_ids)
                pipeline_status["folder_progress"] = {}
                for fid in folder_ids:
                    pipeline_status["folder_progress"][fid] = _init_folder_progress()

        if folder_ids:
            print(f"[PIPELINE] Processing {len(folder_ids)} folder(s): {folder_ids}")

        # Build command
        pipeline_dir = Path(__file__).parent.parent.parent / "master_pipeline"
        cmd = [
            sys.executable,
            str(pipeline_dir / "master_pipeline.py")
        ]

        if download:
            cmd.append("--download")
        if deduplicate:
            cmd.append("--deduplicate")
            if use_llm:
                cmd.append("--use-llm")
            cmd.extend(["--threshold", str(threshold)])
        if biometric:
            cmd.append("--pipeline")
        if folder_ids:
            cmd.extend(["--folder-ids", ",".join(folder_ids)])
        cmd.extend(["--source", source])

        print(f"[PIPELINE] Running command: {' '.join(cmd)}")
        pipeline_status["current_step"] = "initializing"
        pipeline_status["total_folders"] = len(folder_ids) if folder_ids else 1

        # Initialise per-folder tracking
        if folder_ids:
            for fid in folder_ids:
                pipeline_status["folder_progress"][fid] = _init_folder_progress()

        # ── mutable state for closures ──
        state = {"current_folder_id": None, "current_step_name": None}

        def _fp():
            """Shortcut to current folder's progress dict (or global progress)."""
            fid = state["current_folder_id"]
            if fid and fid in pipeline_status.get("folder_progress", {}):
                return pipeline_status["folder_progress"][fid]["steps"]
            return pipeline_status["progress"]

        def _handle_line(line):
            """Process a single line of pipeline output."""
            print(f"[PIPELINE OUTPUT] {line}")

            # ── detect folder change ──
            folder_hdr = re.search(r'FOLDER\s+(\d+)/(\d+):\s*(\S+)', line)
            if folder_hdr:
                idx = int(folder_hdr.group(1))
                total = int(folder_hdr.group(2))
                fid = folder_hdr.group(3)

                # Mark previous folder completed
                prev_fid = state["current_folder_id"]
                if prev_fid and prev_fid in pipeline_status.get("folder_progress", {}):
                    prev = pipeline_status["folder_progress"][prev_fid]
                    if prev["status"] == "running":
                        prev["status"] = "completed"
                        for s in prev["steps"].values():
                            if s["status"] == "running":
                                s["status"] = "completed"

                state["current_folder_id"] = fid
                state["current_step_name"] = None
                pipeline_status["current_folder"] = fid
                pipeline_status["current_folder_idx"] = idx
                pipeline_status["total_folders"] = total

                if fid not in pipeline_status.get("folder_progress", {}):
                    pipeline_status["folder_progress"][fid] = _init_folder_progress()
                pipeline_status["folder_progress"][fid]["status"] = "running"
                return

            # ── detect step change ──
            if "STEP 1:" in line or ("step 1" in line.lower() and ("download" in line.lower() or "sync" in line.lower())):
                state["current_step_name"] = "download"
                pipeline_status["current_step"] = "download"
                _fp()["download"]["status"] = "running"
                _fp()["download"]["message"] = line
                pipeline_status["progress"]["download"]["status"] = "running"
                pipeline_status["progress"]["download"]["message"] = line
            elif "STEP 2:" in line or ("step 2" in line.lower() and "dedup" in line.lower()):
                if _fp()["download"]["status"] == "running":
                    _fp()["download"]["status"] = "completed"
                state["current_step_name"] = "deduplicate"
                pipeline_status["current_step"] = "deduplicate"
                _fp()["deduplicate"]["status"] = "running"
                _fp()["deduplicate"]["message"] = line
                pipeline_status["progress"]["deduplicate"]["status"] = "running"
                pipeline_status["progress"]["deduplicate"]["message"] = line
            elif "STEP 3:" in line or ("step 3" in line.lower() and ("biometric" in line.lower() or "pipeline" in line.lower())):
                if _fp()["deduplicate"]["status"] == "running":
                    _fp()["deduplicate"]["status"] = "completed"
                state["current_step_name"] = "biometric"
                pipeline_status["current_step"] = "biometric"
                _fp()["biometric"]["status"] = "running"
                _fp()["biometric"]["message"] = line
                pipeline_status["progress"]["biometric"]["status"] = "running"
                pipeline_status["progress"]["biometric"]["message"] = line
            elif "STEP 4:" in line or "Consolidate" in line or "Upload Deliverables" in line:
                if _fp()["biometric"]["status"] == "running":
                    _fp()["biometric"]["status"] = "completed"
                state["current_step_name"] = "consolidating"
                pipeline_status["current_step"] = "consolidating"
            elif "PIPELINE COMPLETE" in line:
                fid = state["current_folder_id"]
                if fid and fid in pipeline_status.get("folder_progress", {}):
                    fp = pipeline_status["folder_progress"][fid]
                    fp["status"] = "completed"
                    for s in fp["steps"].values():
                        if s["status"] == "running":
                            s["status"] = "completed"

            # ── parse tqdm / progress numbers  e.g. "Analyzing:  42%|████▏     | 80/192" ──
            progress_match = re.search(r'(\d+)/(\d+)', line)
            cur_step = state["current_step_name"]
            if progress_match and cur_step and cur_step in ("download", "deduplicate", "biometric"):
                current_val = int(progress_match.group(1))
                total_val = int(progress_match.group(2))
                ll = line.lower()

                step = cur_step
                if "download" in ll or "syncing" in ll or "sync" in ll:
                    step = "download"
                elif any(kw in ll for kw in ["compar", "analyz", "duplicat", "dedup", "copying unique", "segregat", "copying"]):
                    step = "deduplicate"
                elif any(kw in ll for kw in ["obfuscat", "stage", "face", "biometric"]):
                    step = "biometric"

                _fp()[step]["current"] = current_val
                _fp()[step]["total"] = total_val
                _fp()[step]["message"] = line

                pipeline_status["progress"][step]["current"] = current_val
                pipeline_status["progress"][step]["total"] = total_val
                pipeline_status["progress"][step]["message"] = line

            # ── detect messages with counts ──
            found_match = re.search(r'Found (\d+) images', line)
            if found_match and cur_step == "download":
                count = int(found_match.group(1))
                _fp()["download"]["total"] = count
                pipeline_status["progress"]["download"]["total"] = count

            if re.search(r'Downloaded (\d+) new images', line):
                _fp()["download"]["message"] = line

            if re.search(r'Found (\d+) duplicate', line):
                _fp()["deduplicate"]["message"] = line

            # ── errors (skip normal summary lines and biometric "verification failed") ──
            is_error_line = "error" in line.lower() or ("failed" in line.lower() and "verification" not in line.lower())
            # "Failed to process: 0" is a summary line meaning zero failures — not an actual error
            if is_error_line and re.search(r'(?:failed|error)[^:]*:\s*0\s*$', line.lower()):
                is_error_line = False
            if is_error_line:
                pipeline_status["errors"].append(line)
                fid = state["current_folder_id"]
                if fid and fid in pipeline_status.get("folder_progress", {}):
                    pipeline_status["folder_progress"][fid]["errors"].append(line)
                    if "❌ Folder" in line or "failed:" in line.lower():
                        pipeline_status["folder_progress"][fid]["status"] = "failed"

        # ── Run pipeline subprocess ──
        import os as _os
        env = {**_os.environ, "PYTHONUNBUFFERED": "1"}
        process = subprocess.Popen(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            bufsize=0,
            cwd=str(pipeline_dir),
            env=env,
        )

        # Read byte-by-byte to handle both \r and \n (tqdm uses \r for progress bars)
        buf = b""
        if process.stdout:
            while True:
                ch = process.stdout.read(1)
                if not ch:
                    if buf:
                        line = buf.decode("utf-8", errors="replace").strip()
                        if line:
                            _handle_line(line)
                    break
                if ch in (b"\n", b"\r"):
                    line = buf.decode("utf-8", errors="replace").strip()
                    buf = b""
                    if line:
                        _handle_line(line)
                else:
                    buf += ch

        # Wait for completion
        returncode = process.wait()
        print(f"[PIPELINE] Process completed with return code: {returncode}")

        # Mark last folder complete
        last_fid = state["current_folder_id"]
        if last_fid and last_fid in pipeline_status.get("folder_progress", {}):
            fp = pipeline_status["folder_progress"][last_fid]
            if fp["status"] == "running":
                fp["status"] = "completed"
                for s in fp["steps"].values():
                    if s["status"] == "running":
                        s["status"] = "completed"

        # Update final status
        pipeline_status["is_running"] = False
        pipeline_status["completed_at"] = datetime.now().isoformat()

        if returncode == 0:
            pipeline_status["current_step"] = "importing"
            for step in pipeline_status["progress"]:
                if pipeline_status["progress"][step]["status"] == "running":
                    pipeline_status["progress"][step]["status"] = "completed"
            print("[PIPELINE] Pipeline completed successfully, now importing to DB...")

            # ── Auto-import images into DB ──
            try:
                backend_dir = Path(__file__).parent.parent.parent
                sys.path.insert(0, str(backend_dir))
                from import_pipeline_images import import_images_from_pipeline
                imported = import_images_from_pipeline()
                print(f"[PIPELINE] Imported {imported} new images into database")
            except Exception as imp_err:
                print(f"[PIPELINE] Auto-import error: {imp_err}")
                import traceback
                traceback.print_exc()
                pipeline_status["errors"].append(f"Auto-import error: {str(imp_err)}")

            # ── Update DriveFolder stats in DB ──
            try:
                _update_drive_folder_stats(folder_ids)
                print("[PIPELINE] Updated DriveFolder stats in DB")
            except Exception as stats_err:
                print(f"[PIPELINE] Stats update error: {stats_err}")
                pipeline_status["errors"].append(f"Stats update error: {str(stats_err)}")

            # ── Cleanup local workspace AFTER stats have been read ──
            try:
                import shutil as _shutil
                bucket_name = os.getenv("GCS_BUCKET_NAME")
                if bucket_name:
                    folders_dir = _get_pipeline_workspace() / "folders"
                    if folders_dir.exists():
                        for fd in sorted(d for d in folders_dir.iterdir() if d.is_dir()):
                            _shutil.rmtree(str(fd), ignore_errors=True)
                        print("[PIPELINE] 🧹 Cleaned up local pipeline workspace (files now in GCS)")
            except Exception as cleanup_err:
                print(f"[PIPELINE] Cleanup warning: {cleanup_err}")

            pipeline_status["current_step"] = "completed"
        else:
            pipeline_status["current_step"] = "failed"
            pipeline_status["errors"].append(f"Pipeline failed with code {returncode}")
            print(f"[PIPELINE] Pipeline failed with code {returncode}")

            # Mark any still-running folders as failed in the DB
            for fid, fp in pipeline_status.get("folder_progress", {}).items():
                if fp["status"] in ("running", "pending"):
                    fp["status"] = "failed"

            # Persist per-folder stats (completed ones stay completed, failed stay failed)
            try:
                _update_drive_folder_stats(folder_ids)
            except Exception:
                pass

    except Exception as e:
        print(f"[PIPELINE] Exception: {str(e)}")
        import traceback
        traceback.print_exc()
        pipeline_status["is_running"] = False
        pipeline_status["current_step"] = "error"
        pipeline_status["completed_at"] = datetime.now().isoformat()
        pipeline_status["errors"].append(f"Exception: {str(e)}")


def reprocess_images_background(image_ids: List[int], db: Session):
    """
    Reprocess specific images in the background.
    
    1. Marks images as pipeline_status='pending' in DB
    2. Removes their files from deliverable/ so biometric step re-processes them
    3. Creates a _reprocess_filenames.json marker so step3 knows to force-reprocess
    4. Re-runs the pipeline for the affected folder(s)
    """
    import subprocess
    import sys
    from pathlib import Path
    
    try:
        images = db.query(Image).filter(Image.id.in_(image_ids)).all()
        
        if not images:
            print("[REPROCESS] No images found for given IDs")
            return
        
        # Group images by folder_id
        folder_to_filenames: Dict[str, list] = {}
        for image in images:
            image.pipeline_status = "pending"
            image.compliance_status = "pending_reprocess"
            fid = image.source_folder_id or "unknown"
            folder_to_filenames.setdefault(fid, []).append(image.filename)
        
        db.commit()
        
        workspace = _get_pipeline_workspace()
        
        # For each folder, remove images from deliverable/ and write reprocess marker
        for fid, filenames in folder_to_filenames.items():
            folder_workspace = workspace / "folders" / fid
            deliverable_dir = folder_workspace / "deliverable"
            
            # Remove from deliverable so biometric step re-processes them
            removed = 0
            for fname in filenames:
                deliverable_file = deliverable_dir / fname
                if deliverable_file.exists():
                    try:
                        deliverable_file.unlink()
                        removed += 1
                    except Exception as e:
                        print(f"[REPROCESS] Could not remove {deliverable_file}: {e}")
            
            # Write reprocess marker file
            marker_path = folder_workspace / "_reprocess_filenames.json"
            try:
                import json
                with open(marker_path, "w") as f:
                    json.dump(filenames, f)
                print(f"[REPROCESS] Folder {fid}: removed {removed} from deliverable, "
                      f"marked {len(filenames)} for reprocess")
            except Exception as e:
                print(f"[REPROCESS] Could not write reprocess marker: {e}")
        
    except Exception as e:
        print(f"Reprocessing error: {e}")
        import traceback
        traceback.print_exc()


@router.post("/sync-status")
def sync_pipeline_status(
    admin: User = Depends(require_admin)
):
    """
    Sync pipeline status from the actual pipeline results file.
    Useful when pipeline was run from terminal instead of UI.
    """
    global pipeline_status
    
    try:
        # Read the actual pipeline results
        backend_dir = Path(__file__).parent.parent.parent
        results_file = backend_dir / "master_pipeline" / "biometric_compliance_pipeline" / "results" / "obfuscation_results.json"
        
        if not results_file.exists():
            raise HTTPException(status_code=404, detail="Pipeline results file not found. Has the pipeline been run?")
        
        import json
        with open(results_file, 'r') as f:
            results = json.load(f)
        
        # Check workspace for actual counts (aggregate across per-folder workspaces)
        workspace = _get_pipeline_workspace()
        folders_dir = workspace / "folders"
        downloaded_count = 0
        unique_count = 0
        
        if folders_dir.exists():
            for fd in folders_dir.iterdir():
                if not fd.is_dir():
                    continue
                dd = fd / "01_downloaded_from_drive"
                dl = fd / "deliverable"
                if dd.exists():
                    downloaded_count += len(list(dd.glob("*.*")))
                if dl.exists():
                    unique_count += len(list(dl.glob("*.*")))
        
        # Fallback: legacy workspace
        if downloaded_count == 0:
            legacy_dd = workspace / "01_downloaded_from_drive"
            legacy_dl = workspace / "deliverable"
            if legacy_dd.exists():
                downloaded_count = len(list(legacy_dd.glob("*.*")))
            if legacy_dl.exists():
                unique_count = len(list(legacy_dl.glob("*.*")))
        
        # Update pipeline status with actual results
        pipeline_status = {
            "is_running": False,
            "current_step": "completed",
            "progress": {
                "download": {
                    "status": "completed",
                    "current": downloaded_count,
                    "total": downloaded_count,
                    "message": f"Downloaded {downloaded_count} images from Google Drive"
                },
                "deduplicate": {
                    "status": "completed",
                    "current": unique_count,
                    "total": unique_count,
                    "message": f"Found {unique_count} unique images"
                },
                "biometric": {
                    "status": "completed",
                    "current": results['total_images'],
                    "total": results['total_images'],
                    "message": f"Processed {results['total_images']} images - {results['statistics']['clean']} clean, {results['statistics']['obfuscated']} obfuscated, {results['statistics']['verification_failed']} QA review"
                }
            },
            "started_at": None,
            "completed_at": datetime.now().isoformat(),
            "errors": [],
            "summary": {
                "total_processed": results['total_images'],
                "clean": results['statistics']['clean'],
                "obfuscated": results['statistics']['obfuscated'],
                "qa_required": results['statistics']['verification_failed'],
                "failed": results['statistics']['failed']
            }
        }
        
        return {
            "success": True,
            "message": "Pipeline status synced successfully",
            "status": pipeline_status
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to sync pipeline status: {str(e)}")


@router.post("/incremental-import")
async def run_incremental_import(
    background_tasks: BackgroundTasks,
    admin: User = Depends(require_admin)
):
    """
    Run incremental pipeline processing.
    
    This intelligently:
    1. Detects NEW downloaded images that haven't been processed
    2. Processes ONLY those new images (not the entire collection)
    3. Imports the newly processed images to the database
    
    Example: If you have images 1-1000 already processed and imported,
    and then download images 1001-1500, this will process only 1001-1500.
    """
    if pipeline_status["is_running"]:
        raise HTTPException(
            status_code=400,
            detail="Pipeline is already running. Please wait for it to complete."
        )
    
    # Run incremental import in background
    background_tasks.add_task(run_incremental_import_background)
    
    return {
        "message": "Incremental import started",
        "info": "Processing only NEW images that haven't been processed yet"
    }


@router.get("/check-new-images")
def check_for_new_images(
    admin: User = Depends(require_admin)
):
    """
    Check how many NEW images are available for processing.
    
    Returns counts of:
    - Downloaded but not processed
    - Processed but not imported to DB
    - Already in database
    """
    try:
        workspace = _get_pipeline_workspace()
        
        # Load state tracking
        state_file = workspace / "pipeline_state.json"
        if state_file.exists():
            with open(state_file, 'r') as f:
                state = json.load(f)
        else:
            state = {"processed_files": [], "imported_files": []}
        
        # Count downloaded files (across per-folder workspaces + legacy)
        extensions = {'.jpg', '.jpeg', '.png', '.webp', '.heic', '.heif', '.avif'}
        downloaded_files = []
        final_files = []
        
        ws_roots = []
        sync_folders_dir = workspace / "folders"
        if sync_folders_dir.is_dir():
            for fd in sorted(sync_folders_dir.iterdir()):
                if fd.is_dir():
                    ws_roots.append(fd)
        ws_roots.append(workspace)  # legacy flat workspace as fallback
        
        seen_dl = set()
        seen_final = set()
        for ws_root in ws_roots:
            dl_dir = ws_root / "01_downloaded_from_drive"
            if dl_dir.exists():
                for f in dl_dir.iterdir():
                    if f.is_file() and f.suffix.lower() in extensions and f.name not in seen_dl:
                        downloaded_files.append(f.name)
                        seen_dl.add(f.name)
            fo_dir = ws_root / "deliverable"
            if fo_dir.exists():
                for f in fo_dir.iterdir():
                    if f.is_file() and f.suffix.lower() in extensions and f.name not in seen_final:
                        final_files.append(f.name)
                        seen_final.add(f.name)
        
        # Count database files
        from app.database import SessionLocal
        db = SessionLocal()
        try:
            db_files = db.execute(text("SELECT filename FROM images")).fetchall()
            db_filenames = {row[0] for row in db_files}
        finally:
            db.close()
        
        # Calculate differences
        processed_set = set(state.get("processed_files", []))
        downloaded_set = set(downloaded_files)
        final_set = set(final_files)
        
        new_downloaded = downloaded_set - processed_set  # Downloaded but not processed
        new_final = final_set - db_filenames  # Processed but not imported
        
        return {
            "new_to_process": len(new_downloaded),
            "new_to_import": len(new_final),
            "total_downloaded": len(downloaded_files),
            "total_processed": len(final_files),
            "total_in_database": len(db_filenames),
            "ready_for_incremental": len(new_downloaded) > 0 or len(new_final) > 0,
            "recommendation": (
                f"Run incremental import to process {len(new_downloaded)} new images and import {len(new_final)} processed images"
                if (len(new_downloaded) > 0 or len(new_final) > 0)
                else "No new images found. All images are up to date."
            )
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error checking for new images: {str(e)}")


# Background task for incremental import
def run_incremental_import_background():
    """Run incremental import in the background."""
    import sys
    global pipeline_status
    
    try:
        pipeline_status["is_running"] = True
        pipeline_status["current_step"] = "incremental_import"
        pipeline_status["started_at"] = datetime.now().isoformat()
        
        backend_dir = Path(__file__).parent.parent.parent
        sys.path.insert(0, str(backend_dir))
        
        # Import and run the incremental importer
        from import_incremental import IncrementalPipelineImporter
        
        importer = IncrementalPipelineImporter()
        results = importer.full_incremental_workflow(
            run_deduplication=False,  # Skip dedup for speed (can be enabled if needed)
            run_biometric=True
        )
        
        pipeline_status["is_running"] = False
        pipeline_status["current_step"] = "completed"
        pipeline_status["completed_at"] = datetime.now().isoformat()
        pipeline_status["summary"] = results.get("summary", {})
        
        print(f"[INCREMENTAL IMPORT] Completed: {results}")
        
    except Exception as e:
        print(f"[INCREMENTAL IMPORT] Error: {str(e)}")
        import traceback
        traceback.print_exc()
        
        pipeline_status["is_running"] = False
        pipeline_status["current_step"] = "error"
        pipeline_status["completed_at"] = datetime.now().isoformat()
        pipeline_status["errors"].append(f"Incremental import error: {str(e)}")


# ── Drive Folder Management ──────────────────────────────────────

@router.get("/folders")
def list_drive_folders(
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin),
):
    """List all registered Google Drive folders with per-folder stats from DB."""
    folders = db.query(DriveFolder).order_by(DriveFolder.added_at.desc()).all()

    # Per-folder image counts from the images table
    folder_image_stats = {}
    rows = db.execute(text("""
        SELECT
            source_folder_id,
            COUNT(*) AS total,
            SUM(CASE WHEN compliance_status IN ('blurred','processed','obfuscated') THEN 1 ELSE 0 END) AS blurred,
            SUM(CASE WHEN compliance_status = 'clean' THEN 1 ELSE 0 END) AS clean,
            SUM(CASE WHEN compliance_status IN ('failed','error') THEN 1 ELSE 0 END) AS failed,
            SUM(CASE WHEN is_improper = TRUE THEN 1 ELSE 0 END) AS improper,
            SUM(CASE WHEN manually_blurred = TRUE THEN 1 ELSE 0 END) AS manually_blurred
        FROM images
        WHERE source_folder_id IS NOT NULL
        GROUP BY source_folder_id
    """)).fetchall()
    for r in rows:
        folder_image_stats[r[0]] = {
            "total_in_db": r[1],
            "blurred": r[2],
            "clean": r[3],
            "failed": r[4],
            "improper": r[5],
            "manually_blurred": r[6],
        }

    result = []
    for f in folders:
        db_stats = folder_image_stats.get(f.folder_id, {})
        result.append({
            "id": f.id,
            "folder_id": f.folder_id,
            "folder_name": f.folder_name or f.folder_id[:12] + "...",
            "status": f.status,
            "added_at": f.added_at.isoformat() if f.added_at else None,
            "last_run_at": f.last_run_at.isoformat() if f.last_run_at else None,
            "notes": f.notes,
            "error_log": f.error_log,
            # Stats from DB images table
            "total_in_db": db_stats.get("total_in_db", 0),
            "blurred": db_stats.get("blurred", 0),
            "clean": db_stats.get("clean", 0),
            "failed": db_stats.get("failed", 0),
            "improper": db_stats.get("improper", 0),
            "manually_blurred": db_stats.get("manually_blurred", 0),
            # Stats from drive_folders table (populated by _update_drive_folder_stats)
            "total_in_drive": f.total_in_drive or 0,
            "unique_in_drive": f.unique_count or 0,
            "duplicates_in_drive": f.duplicate_count or 0,
        })

    # Also count images with no folder_id
    no_folder_count = db.query(Image).filter(
        Image.source_folder_id.is_(None)
    ).count()

    return {
        "folders": result,
        "unassigned_image_count": no_folder_count,
    }


@router.post("/folders/upload-excel")
async def upload_folder_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin),
):
    """
    Upload an Excel (.xlsx/.xls) or CSV file with Google Drive folder IDs.
    
    Expected columns (case-insensitive, flexible naming):
    - folder_id (required): Google Drive folder ID
    - folder_name / name (optional): Human-readable name
    - notes (optional): Any notes about the folder
    """
    import io

    filename = file.filename or ""
    content = await file.read()

    rows = []
    try:
        if filename.endswith(('.xlsx', '.xls')):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            headers_raw = [str(c.value or "").strip().lower().replace(" ", "_") for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {headers_raw[i]: (str(v).strip() if v else None) for i, v in enumerate(row) if i < len(headers_raw)}
                rows.append(row_dict)
        elif filename.endswith('.csv'):
            import csv
            reader = csv.DictReader(io.StringIO(content.decode('utf-8-sig')))
            for row in reader:
                rows.append({k.strip().lower().replace(" ", "_"): (v.strip() if v else None) for k, v in row.items()})
        else:
            raise HTTPException(status_code=400, detail="Unsupported file format. Please upload .xlsx, .xls, or .csv")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")

    if not rows:
        raise HTTPException(status_code=400, detail="File is empty or has no data rows")

    # Normalize column names — look for folder_id column
    fid_key = None
    name_key = None
    notes_key = None
    sample_keys = set()
    for r in rows:
        sample_keys.update(r.keys())
    
    for k in sample_keys:
        kn = k.lower().replace(" ", "_")
        if kn in ("folder_id", "folderid", "drive_folder_id", "google_drive_folder_id", "id"):
            fid_key = k
        elif kn in ("folder_name", "name", "foldername", "label"):
            name_key = k
        elif kn in ("notes", "note", "description", "comment"):
            notes_key = k

    if fid_key is None:
        raise HTTPException(
            status_code=400,
            detail=f"Could not find a 'folder_id' column. Found columns: {list(sample_keys)}"
        )

    added = 0
    skipped = 0
    errors = []
    seen_in_batch = set()  # Track folder IDs seen in this upload to handle duplicates within the file
    
    for idx, row in enumerate(rows, start=2):
        fid = row.get(fid_key)
        if not fid:
            errors.append(f"Row {idx}: missing folder_id")
            continue

        # Skip duplicates within the same file
        if fid in seen_in_batch:
            skipped += 1
            continue
        seen_in_batch.add(fid)

        # Check if already exists in DB
        existing = db.query(DriveFolder).filter(DriveFolder.folder_id == fid).first()
        if existing:
            # Update name/notes if provided
            if name_key and row.get(name_key):
                existing.folder_name = row[name_key]
            if notes_key and row.get(notes_key):
                existing.notes = row[notes_key]
            skipped += 1
            continue

        folder = DriveFolder(
            folder_id=fid,
            folder_name=row.get(name_key) if name_key else None,
            notes=row.get(notes_key) if notes_key else None,
            status="pending",
        )
        db.add(folder)
        added += 1

    db.commit()

    # After initial registration, resolve parent folders into subfolders
    all_new_ids = [row.get(fid_key) for row in rows if row.get(fid_key)]
    all_new_ids = list(dict.fromkeys(all_new_ids))  # dedupe preserving order
    resolved_ids = _resolve_parent_folders(all_new_ids, db)

    return {
        "message": f"Processed {len(rows)} rows: {added} added, {skipped} already existed",
        "added": added,
        "skipped": skipped,
        "errors": errors,
        "total_folders": db.query(DriveFolder).count(),
        "resolved_folders": len(resolved_ids),
    }


@router.post("/folders/add")
def add_drive_folder(
    folder_id: str = Query(...),
    folder_name: str = Query(None),
    notes: str = Query(None),
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin),
):
    """Manually add a single Google Drive folder ID. If it's a parent folder containing
    only subfolders, it will be auto-expanded into child folder entries."""
    existing = db.query(DriveFolder).filter(DriveFolder.folder_id == folder_id).first()
    if existing:
        raise HTTPException(status_code=409, detail="Folder ID already registered")

    # First register it
    folder = DriveFolder(
        folder_id=folder_id,
        folder_name=folder_name,
        notes=notes,
        status="pending",
    )
    db.add(folder)
    db.commit()
    db.refresh(folder)

    # Resolve: if it's a parent, expand into children and remove the parent entry
    resolved = _resolve_parent_folders([folder_id], db)

    if len(resolved) > 1 or (len(resolved) == 1 and resolved[0] != folder_id):
        # Parent was expanded — return the child folders
        child_folders = db.query(DriveFolder).filter(DriveFolder.folder_id.in_(resolved)).all()
        return {
            "message": f"Parent folder expanded into {len(child_folders)} subfolder(s)",
            "expanded": True,
            "folders": [
                {"id": f.id, "folder_id": f.folder_id, "folder_name": f.folder_name, "status": f.status}
                for f in child_folders
            ]
        }

    return {
        "message": "Folder added",
        "folder": {
            "id": folder.id,
            "folder_id": folder.folder_id,
            "folder_name": folder.folder_name,
            "status": folder.status,
        }
    }


@router.delete("/folders/{folder_db_id}")
def remove_drive_folder(
    folder_db_id: int,
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin),
):
    """Remove a Drive folder from the tracking list."""
    folder = db.query(DriveFolder).filter(DriveFolder.id == folder_db_id).first()
    if not folder:
        raise HTTPException(status_code=404, detail="Folder not found")

    db.delete(folder)
    db.commit()
    return {"message": f"Folder '{folder.folder_name or folder.folder_id}' removed"}


@router.get("/folders/stats-table")
def get_folder_stats_table(
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin),
):
    """
    Get per-folder statistics in tabular format.
    Includes both registered folders and images with no folder assignment.
    """
    # All registered folders
    folders = db.query(DriveFolder).order_by(DriveFolder.added_at).all()
    folder_ids = [f.folder_id for f in folders]
    folder_name_map = {f.folder_id: f.folder_name or f.folder_id[:16] + "..." for f in folders}

    # Per-folder image stats from DB
    rows = db.execute(text("""
        SELECT
            COALESCE(source_folder_id, '__unassigned__') AS fid,
            COUNT(*) AS total,
            SUM(CASE WHEN compliance_status IN ('blurred','processed','obfuscated') THEN 1 ELSE 0 END) AS blurred,
            SUM(CASE WHEN compliance_status = 'clean' THEN 1 ELSE 0 END) AS clean,
            SUM(CASE WHEN compliance_status IN ('failed','error') THEN 1 ELSE 0 END) AS failed,
            SUM(CASE WHEN is_improper = TRUE THEN 1 ELSE 0 END) AS improper,
            SUM(CASE WHEN manually_blurred = TRUE THEN 1 ELSE 0 END) AS manually_blurred,
            SUM(CASE WHEN human_faces_detected > 0 THEN 1 ELSE 0 END) AS with_faces,
            SUM(CASE WHEN arbiter_labels IS NOT NULL THEN 1 ELSE 0 END) AS ai_classified
        FROM images
        GROUP BY COALESCE(source_folder_id, '__unassigned__')
    """)).fetchall()

    # Annotation progress per folder (image-level: annotation_status + review_status)
    annotation_rows = db.execute(text("""
        SELECT
            COALESCE(source_folder_id, '__unassigned__') AS fid,
            COUNT(CASE WHEN annotated_by IS NOT NULL THEN 1 END) AS annotated_images,
            COUNT(CASE WHEN annotation_status = 'completed' THEN 1 END) AS completed_annotations,
            COUNT(CASE WHEN review_status = 'approved' THEN 1 END) AS approved_annotations
        FROM images
        GROUP BY COALESCE(source_folder_id, '__unassigned__')
    """)).fetchall()
    ann_map = {r[0]: {"annotated_images": r[1], "completed": r[2], "approved": r[3]} for r in annotation_rows}

    # Build lookup maps from drive_folders DB table
    folder_gcs_count = {f.folder_id: f.total_in_drive or 0 for f in folders}
    folder_error_log = {f.folder_id: f.error_log for f in folders}
    folder_status = {f.folder_id: f.status for f in folders}

    # Build a lookup from the image-stats rows  {fid → row}
    stats_map = {}
    for r in rows:
        stats_map[r[0]] = r

    # Collect all folder IDs to display: registered folders + any with images
    all_fids = list(dict.fromkeys(folder_ids + list(stats_map.keys())))

    table = []
    totals = {"total": 0, "blurred": 0, "clean": 0, "failed": 0, "improper": 0,
              "manually_blurred": 0, "with_faces": 0, "ai_classified": 0,
              "annotated_images": 0, "completed": 0, "approved": 0, "total_in_drive": 0}

    for fid in all_fids:
        r = stats_map.get(fid)
        ann = ann_map.get(fid, {})

        entry = {
            "folder_id": fid if fid != "__unassigned__" else None,
            "folder_name": folder_name_map.get(fid, "Unassigned" if fid == "__unassigned__" else fid[:16] + "..."),
            "total_in_db": r[1] if r else 0,
            "blurred": r[2] if r else 0,
            "clean": r[3] if r else 0,
            "failed": r[4] if r else 0,
            "improper": r[5] if r else 0,
            "manually_blurred": r[6] if r else 0,
            "with_faces": r[7] if r else 0,
            "ai_classified": r[8] if r else 0,
            "annotated_images": ann.get("annotated_images", 0),
            "completed_annotations": ann.get("completed", 0),
            "approved_annotations": ann.get("approved", 0),
            "total_in_drive": folder_gcs_count.get(fid, 0),
            "status": folder_status.get(fid),
            "notes": folder_error_log.get(fid) or None,
        }
        table.append(entry)

        # Accumulate totals
        totals["total"]            += (entry.get("total_in_db") or 0)
        totals["blurred"]          += (entry.get("blurred") or 0)
        totals["clean"]            += (entry.get("clean") or 0)
        totals["failed"]           += (entry.get("failed") or 0)
        totals["improper"]         += (entry.get("improper") or 0)
        totals["manually_blurred"] += (entry.get("manually_blurred") or 0)
        totals["with_faces"]       += (entry.get("with_faces") or 0)
        totals["ai_classified"]    += (entry.get("ai_classified") or 0)
        totals["annotated_images"] += (entry.get("annotated_images") or 0)
        totals["completed"]        += (entry.get("completed_annotations") or 0)
        totals["approved"]         += (entry.get("approved_annotations") or 0)
        totals["total_in_drive"]   += (entry.get("total_in_drive") or 0)

    return {
        "table": table,
        "totals": totals,
    }
