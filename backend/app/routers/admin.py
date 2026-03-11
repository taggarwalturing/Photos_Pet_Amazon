"""
Admin router — rewritten for 3-table schema (users, images, arbiter_predictions).

Tables used : users, images, arbiter_predictions, drive_folders
Categories  : static JSON file (categories.json)
"""
import json
import os
from datetime import datetime, timezone, date, timedelta
from pathlib import Path
from collections import defaultdict

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy import or_, and_, func, cast, Date, case
from sqlalchemy.orm import Session
from typing import Optional

from app.database import get_db
from app.dependencies import require_admin
from app.models.user import User
from app.models.image import Image
from app.schemas.user import UserCreate, UserUpdate, UserResponse
from app.schemas.annotation import (
    ReviewApproveRequest, ReviewUpdateRequest,
    ReviewTableCategoryCell, ReviewTableRow, ReviewTableCategory, ReviewTableResponse,
    ReworkRequest, ImageReworkRequest,
)
from app.services.auth import hash_password
from app.utils.deliverable import check_and_deliver_image, update_deliverable_if_delivered
from app.utils import categories as categories_util
from pydantic import BaseModel


router = APIRouter(prefix="/admin", tags=["Admin"])


# ── User Management ──────────────────────────────────────────────

@router.get("/users", response_model=list[UserResponse])
def list_users(
    db: Session = Depends(get_db),
    _admin: User = Depends(require_admin),
):
    users = db.query(User).order_by(User.id).all()
    total_images = db.query(Image).count()

    # Get today's date for daily stats
    today = date.today()

    result = []
    for u in users:
        # Count images this user annotated (completed)
        completed_annotations = (
            db.query(Image)
            .filter(Image.annotated_by == u.id, Image.annotation_status == "completed")
            .count()
        )
        
        # Count images annotated today
        today_image_count = (
            db.query(Image)
            .filter(
                Image.annotated_by == u.id,
                Image.annotation_status == "completed",
                cast(Image.annotated_at, Date) == today,
            )
            .count()
        )

        # Count images actually assigned to this user
        actual_assigned = (
            db.query(Image)
            .filter(Image.assigned_to == u.id)
            .count()
        )
        
        result.append(UserResponse(
            id=u.id,
            username=u.username,
            full_name=u.full_name,
            role=u.role,
            is_active=u.is_active,
            created_at=u.created_at,
            total_images=total_images if u.role == "annotator" else 0,
            assigned_image_count=u.assigned_image_count or 0,
            completed_annotations=completed_annotations,
            today_image_count=today_image_count,
            actual_assigned=actual_assigned,
        ))
    return result


@router.get("/users/daily-stats")
def get_daily_annotation_stats(
    days: int = Query(7, ge=1, le=90),
    db: Session = Depends(get_db),
    _admin: User = Depends(require_admin),
):
    """
    Get daily annotation counts per annotator for the last N days.
    Returns how many distinct images each annotator completed per day.
    """
    start_date = date.today() - timedelta(days=days - 1)

    # Query: group by annotated_by and date, count distinct images
    rows = (
        db.query(
            Image.annotated_by,
            cast(Image.annotated_at, Date).label("annotation_date"),
            func.count(Image.id).label("image_count"),
        )
        .filter(
            Image.annotation_status == "completed",
            Image.annotated_at.isnot(None),
            cast(Image.annotated_at, Date) >= start_date,
        )
        .group_by(Image.annotated_by, cast(Image.annotated_at, Date))
        .all()
    )

    # Get annotator usernames
    annotators = db.query(User).filter(User.role == "annotator").all()
    username_map = {u.id: u.username for u in annotators}

    # Build per-annotator daily stats
    stats = {}
    for annotator_id, annotation_date, image_count in rows:
        username = username_map.get(annotator_id, f"user_{annotator_id}")
        if username not in stats:
            stats[username] = {"annotator_id": annotator_id, "daily": {}}
        stats[username]["daily"][str(annotation_date)] = image_count

    # Build date range
    date_range = [str(start_date + timedelta(days=i)) for i in range(days)]

    return {
        "date_range": date_range,
        "annotators": stats,
    }


@router.get("/annotator-stats")
def get_annotator_stats(
    days: int = Query(7, ge=1, le=90),
    db: Session = Depends(get_db),
    _admin: User = Depends(require_admin),
):
    """
    Comprehensive daily stats per annotator:
      - annotated: distinct images with completed annotations per day
      - blurred: images blurred by annotator per day (manually_blurred_at)
      - restored: images restored by annotator per day (is_restore_annotator)
      - approved: annotator's images approved per day (reviewed_at)
    """
    start_date = date.today() - timedelta(days=days - 1)

    # ── All annotators ──
    annotators = db.query(User).filter(User.role == "annotator").all()
    id_to_name = {u.id: u.username for u in annotators}

    # Prepare result skeleton per annotator
    result = {}
    for u in annotators:
        result[u.username] = {
            "annotator_id": u.id,
            "annotated": {},
            "blurred": {},
            "restored": {},
            "approved": {},
            "totals": {"annotated": 0, "blurred": 0, "restored": 0, "approved": 0},
        }

    # ── 1. Annotated per day (images completed) ──
    annotated_rows = (
        db.query(
            Image.annotated_by,
            cast(Image.annotated_at, Date).label("d"),
            func.count(Image.id).label("cnt"),
        )
        .filter(
            Image.annotation_status == "completed",
            Image.annotated_at.isnot(None),
            cast(Image.annotated_at, Date) >= start_date,
        )
        .group_by(Image.annotated_by, cast(Image.annotated_at, Date))
        .all()
    )
    for annotator_id, d, cnt in annotated_rows:
        name = id_to_name.get(annotator_id)
        if name and name in result:
            result[name]["annotated"][str(d)] = cnt

    # ── 2. Blurred per day (images where manually_blurred_by = annotator) ──
    blurred_rows = (
        db.query(
            Image.manually_blurred_by,
            cast(Image.manually_blurred_at, Date).label("d"),
            func.count(Image.id).label("cnt"),
        )
        .filter(
            Image.manually_blurred_by.isnot(None),
            Image.manually_blurred_at.isnot(None),
            cast(Image.manually_blurred_at, Date) >= start_date,
        )
        .group_by(Image.manually_blurred_by, cast(Image.manually_blurred_at, Date))
        .all()
    )
    for user_id, d, cnt in blurred_rows:
        name = id_to_name.get(user_id)
        if name and name in result:
            result[name]["blurred"][str(d)] = cnt

    # ── 3. Restored per day ──
    # In the new schema, is_restore_annotator is a boolean flag.
    # We can approximate restored count using images where is_restore_annotator=True
    # but there's no per-event timestamp, so we only show totals.

    # ── 4. Approved per day (annotator's images that were approved, grouped by reviewed_at) ──
    approved_rows = (
        db.query(
            Image.annotated_by,
            cast(Image.reviewed_at, Date).label("d"),
            func.count(Image.id).label("cnt"),
        )
        .filter(
            Image.review_status == "approved",
            Image.reviewed_at.isnot(None),
            cast(Image.reviewed_at, Date) >= start_date,
        )
        .group_by(Image.annotated_by, cast(Image.reviewed_at, Date))
        .all()
    )
    for annotator_id, d, cnt in approved_rows:
        name = id_to_name.get(annotator_id)
        if name and name in result:
            result[name]["approved"][str(d)] = cnt

    # ── Compute cumulative totals ──
    for u in annotators:
        name = u.username
        if name not in result:
            continue
        result[name]["totals"]["annotated"] = (
            db.query(func.count(Image.id))
            .filter(Image.annotated_by == u.id, Image.annotation_status == "completed")
            .scalar() or 0
        )
        result[name]["totals"]["blurred"] = (
            db.query(func.count(Image.id))
            .filter(Image.manually_blurred_by == u.id)
            .scalar() or 0
        )
        result[name]["totals"]["restored"] = (
            db.query(func.count(Image.id))
            .filter(Image.is_restore_annotator == True, Image.annotated_by == u.id)
            .scalar() or 0
        )
        result[name]["totals"]["approved"] = (
            db.query(func.count(Image.id))
            .filter(Image.annotated_by == u.id, Image.review_status == "approved")
            .scalar() or 0
        )

    # Build date range
    date_range = [str(start_date + timedelta(days=i)) for i in range(days)]

    return {
        "date_range": date_range,
        "annotators": result,
    }


@router.post("/users", response_model=UserResponse, status_code=status.HTTP_201_CREATED)
def create_user(
    payload: UserCreate,
    db: Session = Depends(get_db),
    _admin: User = Depends(require_admin),
):
    existing = db.query(User).filter(User.username == payload.username).first()
    if existing:
        raise HTTPException(status_code=400, detail="Username already exists")
    user = User(
        username=payload.username,
        password_hash=hash_password(payload.password),
        full_name=payload.full_name,
        role=payload.role,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return UserResponse(
        id=user.id,
        username=user.username,
        full_name=user.full_name,
        role=user.role,
        is_active=user.is_active,
        created_at=user.created_at,
    )


@router.put("/users/{user_id}", response_model=UserResponse)
def update_user(
    user_id: int,
    payload: UserUpdate,
    db: Session = Depends(get_db),
    _admin: User = Depends(require_admin),
):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if payload.full_name is not None:
        user.full_name = payload.full_name
    if payload.is_active is not None:
        user.is_active = payload.is_active
    if payload.password is not None:
        user.password_hash = hash_password(payload.password)
    if payload.assigned_image_count is not None:
        user.assigned_image_count = payload.assigned_image_count
    db.commit()
    db.refresh(user)
    actual_assigned = db.query(Image).filter(Image.assigned_to == user.id).count()
    return UserResponse(
        id=user.id,
        username=user.username,
        full_name=user.full_name,
        role=user.role,
        is_active=user.is_active,
        created_at=user.created_at,
        assigned_image_count=user.assigned_image_count or 0,
        actual_assigned=actual_assigned,
    )


@router.post("/assign-images")
def assign_images_to_annotators(
    db: Session = Depends(get_db),
    _admin: User = Depends(require_admin),
):
    """
    Assign images to annotators in continuous sequential blocks based on
    each annotator's `assigned_image_count`.

    Algorithm:
    1. Get all active annotators (role=annotator, is_active=True) ordered by id.
    2. Get all assignable images (non-duplicate, ordered by id).
    3. Clear all existing assignments.
    4. For each annotator, assign the next N images in sequence.

    Only annotators with assigned_image_count > 0 get images.
    Images beyond the total assigned count remain unassigned (invisible to all annotators).
    """
    # Get active annotators ordered by ID, only those with an assignment count > 0
    annotators = (
        db.query(User)
        .filter(
            User.role == "annotator",
            User.is_active == True,  # noqa: E712
            User.assigned_image_count > 0,
        )
        .order_by(User.id)
        .all()
    )

    # Get all assignable images (non-duplicate) ordered by ID
    all_images = (
        db.query(Image)
        .filter(Image.is_duplicate == False)  # noqa: E712
        .order_by(Image.id)
        .all()
    )

    # Clear all existing assignments first
    db.query(Image).filter(Image.assigned_to.isnot(None)).update(
        {"assigned_to": None}, synchronize_session="fetch"
    )

    # Assign in sequential blocks
    cursor = 0
    assignment_summary = []
    for annotator in annotators:
        count = annotator.assigned_image_count or 0
        if count <= 0:
            continue

        # Slice the next `count` images for this annotator
        end = min(cursor + count, len(all_images))
        assigned_ids = []
        for i in range(cursor, end):
            all_images[i].assigned_to = annotator.id
            assigned_ids.append(all_images[i].id)

        assignment_summary.append({
            "annotator_id": annotator.id,
            "username": annotator.username,
            "requested": count,
            "assigned": len(assigned_ids),
            "image_id_range": f"{assigned_ids[0]}–{assigned_ids[-1]}" if assigned_ids else "none",
        })

        cursor = end

    db.commit()

    unassigned_count = len(all_images) - cursor
    return {
        "total_images": len(all_images),
        "total_assigned": cursor,
        "unassigned": unassigned_count,
        "assignments": assignment_summary,
    }


@router.get("/assignment-summary")
def get_assignment_summary(
    db: Session = Depends(get_db),
    _admin: User = Depends(require_admin),
):
    """Get current image assignment summary for all annotators."""
    annotators = (
        db.query(User)
        .filter(User.role == "annotator")
        .order_by(User.id)
        .all()
    )

    total_images = db.query(Image).filter(Image.is_duplicate == False).count()  # noqa: E712
    total_assigned = db.query(Image).filter(Image.assigned_to.isnot(None)).count()

    summary = []
    for a in annotators:
        actual = db.query(Image).filter(Image.assigned_to == a.id).count()
        completed = (
            db.query(Image)
            .filter(Image.assigned_to == a.id, Image.annotation_status == "completed")
            .count()
        )
        summary.append({
            "annotator_id": a.id,
            "username": a.username,
            "requested": a.assigned_image_count or 0,
            "actual_assigned": actual,
            "completed": completed,
            "pending": actual - completed,
        })

    return {
        "total_images": total_images,
        "total_assigned": total_assigned,
        "unassigned": total_images - total_assigned,
        "annotators": summary,
    }


# ── Categories ────────────────────────────────────────────────────

@router.get("/categories")
def list_categories(
    _admin: User = Depends(require_admin),
):
    """Return categories from static JSON (no DB query needed)."""
    all_cats = categories_util.get_categories()
    return [
        {
            "id": c["id"],
            "name": c["name"],
            "key": c["key"],
            "display_order": c["display_order"],
            "options": c["options"],
        }
        for c in all_cats
    ]


# ── Images ────────────────────────────────────────────────────────

@router.get("/images")
def list_images(
    db: Session = Depends(get_db),
    _admin: User = Depends(require_admin),
    status_filter: Optional[str] = Query(None, alias="filter"),
    search: Optional[str] = Query(None),
):
    query = db.query(Image).order_by(Image.id)

    # Apply search
    if search:
        query = query.filter(Image.filename.ilike(f"%{search}%"))

    # Apply filter
    if status_filter == "blurred":
        query = query.filter(
            or_(
                Image.compliance_status.in_(["blurred", "processed", "obfuscated"]),
                Image.manually_blurred == True,
            )
        )
    elif status_filter == "clean":
        query = query.filter(
            Image.compliance_status == "clean",
            or_(Image.manually_blurred == False, Image.manually_blurred.is_(None)),
        )
    elif status_filter == "manually_blurred":
        query = query.filter(Image.manually_blurred == True)
    elif status_filter == "ai_generated":
        query = query.filter(Image.is_ai_generated == True)
    elif status_filter == "human_visible":
        query = query.filter(Image.human_visible == True)
    elif status_filter == "improper":
        query = query.filter(Image.is_improper == True)
    elif status_filter == "delivered":
        query = query.filter(Image.deliverable_image_path.isnot(None))
    elif status_filter == "not_delivered":
        query = query.filter(Image.deliverable_image_path.is_(None))

    images = query.all()

    # ── Summary stats (unfiltered, single aggregate query) ──
    summary_row = db.query(
        func.count().label("total"),
        func.sum(case(
            (or_(
                Image.compliance_status.in_(["blurred", "processed", "obfuscated"]),
                Image.manually_blurred == True,
            ), 1), else_=0
        )).label("blurred"),
        func.sum(case(
            (and_(
                Image.compliance_status == "clean",
                or_(Image.manually_blurred == False, Image.manually_blurred.is_(None)),
            ), 1), else_=0
        )).label("clean"),
        func.sum(case((Image.manually_blurred == True, 1), else_=0)).label("manually_blurred"),
        func.sum(case((Image.is_ai_generated == True, 1), else_=0)).label("ai_generated"),
        func.sum(case((Image.human_visible == True, 1), else_=0)).label("human_visible"),
        func.sum(case((Image.is_improper == True, 1), else_=0)).label("improper"),
        func.sum(case((Image.deliverable_image_path.isnot(None), 1), else_=0)).label("delivered"),
    ).one()

    # ── Load all categories from static JSON ──
    all_categories = categories_util.get_categories()
    cat_by_key = {c["key"]: c for c in all_categories}

    # Build option label lookups from static JSON
    option_label_map = {}  # option_id → label
    for cat in all_categories:
        for o in cat["options"]:
            option_label_map[o["id"]] = o["label"]

    # Build per-image label data from Image.annotations JSON
    def _build_labels(img):
        """Resolve labels: human annotations take priority, then AI predictions."""
        category_labels = {}
        label_source = {}
        annotation_status = {}

        image_annotations = img.annotations or {}
        arbiter_labels = img.arbiter_labels or {}

        for cat in all_categories:
            cat_key = cat["key"]
            cat_id_str = str(cat["id"])

            # Check human annotations
            ann_data = image_annotations.get(cat_key, {})
            selected_option_ids = ann_data.get("selected_option_ids", [])

            if selected_option_ids:
                # Human annotation exists — use stored labels if available, else resolve IDs
                labels = ann_data.get("selected_labels")
                if not labels:
                    labels = []
                    for opt_id in selected_option_ids:
                        opt = categories_util.get_option_by_id(opt_id)
                        if opt:
                            labels.append(opt["label"])
                category_labels[cat_id_str] = labels

                if img.review_status == "approved":
                    label_source[cat_id_str] = "approved"
                    annotation_status[cat_id_str] = "approved"
                elif img.review_status == "rework_requested":
                    label_source[cat_id_str] = "rework"
                    annotation_status[cat_id_str] = "rework"
                else:
                    label_source[cat_id_str] = "human"
                    annotation_status[cat_id_str] = "completed"
            elif arbiter_labels.get(cat_key):
                # AI prediction exists — use stored label directly if available
                pred_data = arbiter_labels[cat_key]
                option_label = pred_data.get("label") if isinstance(pred_data, dict) else None
                if not option_label:
                    pred_key = (
                        pred_data.get("final") or pred_data.get("key")
                        if isinstance(pred_data, dict)
                        else str(pred_data) if pred_data else None
                    )
                    if pred_key:
                        option_label = categories_util.arbiter_label_to_option_label(pred_key)
                if option_label:
                    category_labels[cat_id_str] = [option_label]
                    label_source[cat_id_str] = "ai"
                    annotation_status[cat_id_str] = "ai_predicted"

        return category_labels, label_source, annotation_status

    images_out = []
    for img in images:
        cat_labels, lbl_source, ann_status = _build_labels(img)
        images_out.append({
            "id": img.id,
            "filename": img.filename,
            "original_filename": img.original_filename,
            "url": img.url,
            "created_at": img.created_at,
            "compliance_status": img.compliance_status,
            "manually_blurred": img.manually_blurred or False,
            "is_ai_generated": img.is_ai_generated or False,
            "human_visible": img.human_visible,
            "is_improper": img.is_improper or False,
            "human_faces_detected": img.human_faces_detected or 0,
            "is_using_processed": img.is_using_processed if img.is_using_processed is not None else True,
            "is_blurred": (img.manually_blurred or False) or (
                (img.is_using_processed is not False) and
                (img.compliance_status or "") in ("blurred", "processed", "obfuscated")
            ),
            "source_folder_id": img.source_folder_id,
            "image_drive_id": img.image_drive_id,
            "is_blurred_annotator": img.is_blurred_annotator or False,
            "is_restore_annotator": img.is_restore_annotator or False,
            "deliverable_image_path": img.deliverable_image_path,
            "is_manually_modified": img.is_manually_modified,
            # Label data
            "category_labels": cat_labels,
            "category_label_source": lbl_source,
            "annotation_status": ann_status,
        })

    return {
        "summary": {
            "total": summary_row.total or 0,
            "blurred": summary_row.blurred or 0,
            "clean": summary_row.clean or 0,
            "manually_blurred": summary_row.manually_blurred or 0,
            "ai_generated": summary_row.ai_generated or 0,
            "human_visible": summary_row.human_visible or 0,
            "improper": summary_row.improper or 0,
            "delivered": summary_row.delivered or 0,
        },
        "total": len(images_out),
        "categories": [
            {"id": c["id"], "name": c["name"]}
            for c in all_categories
        ],
        "images": images_out,
    }


# ── Single image status (for admin lightbox) ─────────────────────

@router.get("/images/{image_id}/status")
def get_image_status(
    image_id: int,
    db: Session = Depends(get_db),
    _admin: User = Depends(require_admin),
):
    """Return blur-related status flags for a single image (admin-accessible)."""
    img = db.query(Image).filter(Image.id == image_id).first()
    if not img:
        raise HTTPException(status_code=404, detail="Image not found")

    _manually_blurred = img.manually_blurred or False
    _is_using_processed = img.is_using_processed if img.is_using_processed is not None else True
    _compliance_status = img.compliance_status or None
    _is_blurred = _manually_blurred or (
        (img.is_using_processed is not False) and
        (_compliance_status or "") in ("blurred", "processed", "obfuscated")
    )

    return {
        "id": img.id,
        "is_blurred": _is_blurred,
        "compliance_status": _compliance_status,
        "is_using_processed": _is_using_processed,
        "manually_blurred": _manually_blurred,
        "is_blurred_annotator": img.is_blurred_annotator or False,
        "is_restore_annotator": img.is_restore_annotator or False,
        "deliverable_image_path": img.deliverable_image_path,
        "is_manually_modified": img.is_manually_modified,
    }


# ── Review ────────────────────────────────────────────────────────

@router.get("/review/table", response_model=ReviewTableResponse)
def review_table(
    annotator_id: Optional[int] = Query(None),
    review_status: Optional[str] = Query(None),  # pending, approved
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin),
):
    """
    Spreadsheet-style view: images as rows, categories as columns.
    Uses Image.annotations JSON and Image.review_status for review workflow.
    """
    # Base query: images that have been annotated
    base_q = db.query(Image).filter(
        or_(
            Image.annotation_status == "completed",
            and_(Image.annotation_status == "in_progress", Image.review_status == "rework_requested"),
        )
    )
    if annotator_id is not None:
        base_q = base_q.filter(Image.annotated_by == annotator_id)
    if review_status == "pending":
        base_q = base_q.filter(
            or_(
                Image.review_status.is_(None),
                Image.review_status == "pending",
                Image.review_status == "rework_requested",
                Image.review_status == "rework_completed",
            )
        )
    elif review_status == "approved":
        base_q = base_q.filter(Image.review_status == "approved")

    # Get total count BEFORE pagination
    total_images = base_q.count()

    # Paginate
    images = (
        base_q.order_by(Image.id)
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    # Load categories from static JSON
    all_categories = categories_util.get_categories()

    # Build option lookups
    option_map = {}  # option_id → {id, label, is_typical}
    for cat in all_categories:
        for o in cat["options"]:
            option_map[o["id"]] = o

    # Build rows
    rows = []
    for img in images:
        image_annotations = img.annotations or {}
        arbiter_labels = img.arbiter_labels or {}

        # Get annotator username
        annotated_by_username = None
        if img.annotator:
            annotated_by_username = img.annotator.username

        # Get reviewer username
        reviewed_by_username = None
        if img.reviewed_by:
            reviewer = db.query(User).filter(User.id == img.reviewed_by).first()
            if reviewer:
                reviewed_by_username = reviewer.username

        # Build per-category cells
        category_cells = {}
        for cat in all_categories:
            cat_key = cat["key"]
            cat_id_str = str(cat["id"])

            ann_data = image_annotations.get(cat_key, {})
            selected_option_ids = ann_data.get("selected_option_ids", [])

            # Resolve selected option labels — use stored labels if available, else resolve IDs
            selected_options = []
            stored_labels = ann_data.get("selected_labels", [])
            if stored_labels and len(stored_labels) == len(selected_option_ids):
                for oid, lbl in zip(selected_option_ids, stored_labels):
                    selected_options.append({"id": oid, "label": lbl})
            else:
                for opt_id in selected_option_ids:
                    opt = categories_util.get_option_by_id(opt_id)
                    if opt:
                        selected_options.append({"id": opt["id"], "label": opt["label"]})

            # All options for this category
            all_options = [
                {"id": o["id"], "label": o["label"], "is_typical": o["is_typical"]}
                for o in sorted(cat["options"], key=lambda x: x["display_order"])
            ]

            # Determine label source
            if selected_option_ids:
                if img.review_status == "approved":
                    source = "approved"
                elif img.review_status == "rework_requested":
                    source = "rework"
                else:
                    source = "human"
            elif arbiter_labels.get(cat_key):
                source = "ai"
                pred_data = arbiter_labels[cat_key]
                option_label = pred_data.get("label") if isinstance(pred_data, dict) else None
                if not option_label:
                    pred_key = (
                        pred_data.get("final") or pred_data.get("key")
                        if isinstance(pred_data, dict)
                        else str(pred_data) if pred_data else None
                    )
                    if pred_key:
                        option_label = categories_util.arbiter_label_to_option_label(pred_key)
                if option_label:
                    selected_options = [{"id": 0, "label": option_label}]
            else:
                source = "pending"

            category_cells[cat_id_str] = ReviewTableCategoryCell(
                selected_options=selected_options,
                all_options=all_options,
                label_source=source,
            )

        # Compute is_blurred
        _manually_blurred = img.manually_blurred or False
        _is_blurred = _manually_blurred or (
            (img.is_using_processed is not False) and
            (img.compliance_status or "") in ("blurred", "processed", "obfuscated")
        )

        rows.append(ReviewTableRow(
            image_id=img.id,
            image_url=img.url or "",
            image_filename=img.filename,
            annotation_status=img.annotation_status or "pending",
            annotations=category_cells,
            annotated_by_username=annotated_by_username,
            review_status=img.review_status,
            review_note=img.review_note,
            reviewed_by_username=reviewed_by_username,
            reviewed_at=img.reviewed_at,
            is_blurred=_is_blurred,
            compliance_status=img.compliance_status,
            manually_blurred=_manually_blurred,
            is_blurred_annotator=img.is_blurred_annotator or False,
            is_restore_annotator=img.is_restore_annotator or False,
            deliverable_image_path=img.deliverable_image_path,
            is_manually_modified=img.is_manually_modified,
            gcs_folder=img.gcs_folder or "input",
        ))

    # Category headers
    cat_list = [ReviewTableCategory(id=c["id"], name=c["name"]) for c in all_categories]

    return ReviewTableResponse(
        images=rows,
        categories=cat_list,
        total_images=total_images,
        page=page,
        page_size=page_size,
    )


@router.get("/review/stats")
def review_stats(
    db: Session = Depends(get_db),
    _admin: User = Depends(require_admin),
):
    """Get image-level review statistics."""
    # Images with completed annotations
    total_annotated = (
        db.query(Image)
        .filter(Image.annotation_status == "completed")
        .count()
    )

    approved_images = (
        db.query(Image)
        .filter(Image.review_status == "approved")
        .count()
    )

    pending_images = (
        db.query(Image)
        .filter(
            Image.annotation_status == "completed",
            or_(
                Image.review_status.is_(None),
                Image.review_status == "pending",
                Image.review_status == "rework_completed",
            ),
        )
        .count()
    )

    rework_images = (
        db.query(Image)
        .filter(Image.review_status == "rework_requested")
        .count()
    )

    # Deliverable stats
    delivered = db.query(Image).filter(Image.deliverable_image_path.isnot(None)).count()
    delivered_modified = db.query(Image).filter(
        Image.deliverable_image_path.isnot(None),
        Image.is_manually_modified == True,
    ).count()
    delivered_original = db.query(Image).filter(
        Image.deliverable_image_path.isnot(None),
        or_(Image.is_manually_modified == False, Image.is_manually_modified.is_(None)),
    ).count()

    return {
        "total_completed": total_annotated,
        "pending_review": pending_images,
        "rework_requested": rework_images,
        "approved": approved_images,
        "delivered": delivered,
        "delivered_modified": delivered_modified,
        "delivered_original": delivered_original,
    }


@router.get("/deliverable/stats")
def deliverable_stats(
    db: Session = Depends(get_db),
    _admin: User = Depends(require_admin),
):
    """Get deliverable image statistics grouped by folder_id."""
    delivered_images = (
        db.query(Image)
        .filter(Image.deliverable_image_path.isnot(None))
        .all()
    )

    # Group by folder_id
    folder_stats = {}
    for img in delivered_images:
        fid = img.source_folder_id or "unknown"
        if fid not in folder_stats:
            folder_stats[fid] = {"folder_id": fid, "total": 0, "modified": 0, "original": 0}
        folder_stats[fid]["total"] += 1
        if img.is_manually_modified:
            folder_stats[fid]["modified"] += 1
        else:
            folder_stats[fid]["original"] += 1

    total_delivered = len(delivered_images)
    total_images = db.query(Image).count()

    return {
        "total_images": total_images,
        "total_delivered": total_delivered,
        "total_pending": total_images - total_delivered,
        "by_folder": list(folder_stats.values()),
    }


# ── Approve / Update Image (image-level review) ──────────────────

@router.put("/review/image/{image_id}/approve")
def approve_image(
    image_id: int,
    payload: ReviewApproveRequest,
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin),
):
    """Approve an image's annotations. If approved, copy to deliverable."""
    image = db.query(Image).filter(Image.id == image_id).first()
    if not image:
        raise HTTPException(status_code=404, detail="Image not found")

    if not image.annotations or image.annotation_status != "completed":
        raise HTTPException(status_code=400, detail="Image has no completed annotations to approve")

    # ── Log approval in history ──
    now = datetime.now(timezone.utc)
    history = list(image.annotation_history or [])
    history.append({
        "ts": now.isoformat(),
        "by": admin.id,
        "username": admin.username,
        "role": "reviewer",
        "action": "approve",
        "review_note": payload.review_note,
    })
    image.annotation_history = history

    image.review_status = "approved"
    image.review_note = payload.review_note
    image.reviewed_by = admin.id
    image.reviewed_at = now
    db.commit()

    # Copy to deliverable
    check_and_deliver_image(image_id, db)

    return {"message": "Image approved", "image_id": image_id}


@router.put("/review/image/{image_id}/update")
def update_and_approve_image(
    image_id: int,
    payload: ReviewUpdateRequest,
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin),
):
    """Admin edits the label selections and approves the image."""
    image = db.query(Image).filter(Image.id == image_id).first()
    if not image:
        raise HTTPException(status_code=404, detail="Image not found")

    # ── Append to annotation_history before overwriting ──
    now = datetime.now(timezone.utc)
    history = list(image.annotation_history or [])
    history.append({
        "ts": now.isoformat(),
        "by": admin.id,
        "username": admin.username,
        "role": "reviewer",
        "action": "edit",
        "annotations": payload.annotations,
        "review_note": payload.review_note or "Edited by admin",
    })
    image.annotation_history = history

    # Enrich annotations with human-readable labels and save
    image.annotations = categories_util.enrich_annotations_with_labels(payload.annotations)
    image.annotation_status = "completed"
    image.review_status = "approved"
    image.review_note = payload.review_note or "Edited by admin"
    image.reviewed_by = admin.id
    image.reviewed_at = now
    db.commit()

    # Copy to deliverable
    check_and_deliver_image(image_id, db)

    return {"message": "Image updated and approved", "image_id": image_id}


# ── Send for Rework ──────────────────────────────────────────────

@router.post("/images/{image_id}/rework")
def send_image_for_rework(
    image_id: int,
    payload: ImageReworkRequest,
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin),
):
    """
    Send an image back to the annotator for rework.
    Resets the annotation status and sets review_status to rework_requested.
    """
    image = db.query(Image).filter(Image.id == image_id).first()
    if not image:
        raise HTTPException(status_code=404, detail="Image not found")

    if not image.annotated_by:
        raise HTTPException(status_code=400, detail="Image has not been annotated yet")

    now = datetime.now(timezone.utc)

    # ── Log rework request in history ──
    history = list(image.annotation_history or [])
    history.append({
        "ts": now.isoformat(),
        "by": admin.id,
        "username": admin.username,
        "role": "reviewer",
        "action": "rework",
        "review_note": payload.reason,
    })
    image.annotation_history = history

    image.annotation_status = "in_progress"
    image.review_status = "rework_requested"
    image.review_note = payload.reason
    image.reviewed_by = admin.id
    image.reviewed_at = now

    db.commit()

    return {
        "message": "Image sent for rework",
        "image_id": image_id,
        "annotator_id": image.annotated_by,
    }


# Backward compatibility — accept annotation_id but treat as image-level rework
@router.post("/annotations/{annotation_id}/rework")
def send_annotation_for_rework_compat(
    annotation_id: int,
    payload: ReworkRequest,
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin),
):
    """
    Legacy endpoint — accepts annotation_id but the new schema is image-level.
    The annotation_id is actually the image_id in the new schema.
    """
    return send_image_for_rework(
        image_id=annotation_id,
        payload=ImageReworkRequest(reason=payload.reason),
        db=db,
        admin=admin,
    )


# ── Improper Images ───────────────────────────────────────────────

@router.get("/images/improper")
def list_improper_images(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    _admin: User = Depends(require_admin),
):
    """List all images marked as improper by annotators."""
    query = (
        db.query(Image)
        .filter(Image.is_improper == True)
        .order_by(Image.marked_improper_at.desc())
    )
    
    total = query.count()
    images = query.offset((page - 1) * page_size).limit(page_size).all()
    
    result = []
    for img in images:
        marker = None
        if img.marked_improper_by:
            marker = db.query(User).filter(User.id == img.marked_improper_by).first()
        
        result.append({
            "id": img.id,
            "filename": img.filename,
            "url": img.url,
            "is_improper": img.is_improper,
            "improper_reason": img.improper_reason,
            "marked_improper_by": marker.username if marker else None,
            "marked_improper_by_id": img.marked_improper_by,
            "marked_improper_at": img.marked_improper_at,
        })
    
    return {
        "images": result,
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@router.get("/images/improper/count")
def get_improper_images_count(
    db: Session = Depends(get_db),
    _admin: User = Depends(require_admin),
):
    """Get count of improper images."""
    count = db.query(Image).filter(Image.is_improper == True).count()
    return {"count": count}


@router.put("/images/{image_id}/revoke-improper")
def revoke_improper_status(
    image_id: int,
    db: Session = Depends(get_db),
    _admin: User = Depends(require_admin),
):
    """Admin revokes improper status - marks image as proper again."""
    image = db.query(Image).filter(Image.id == image_id).first()
    if not image:
        raise HTTPException(status_code=404, detail="Image not found")
    
    if not image.is_improper:
        raise HTTPException(status_code=400, detail="Image is not marked as improper")
    
    image.is_improper = False
    image.improper_reason = None
    image.marked_improper_by = None
    image.marked_improper_at = None
    
    db.commit()
    
    return {
        "message": "Image marked as proper again",
        "image_id": image_id,
    }


# ── Photo Registry (All Downloaded Photos Status) ──────────────────

@router.get("/photo-registry")
def get_photo_registry(
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=10, le=200),
    search: str = Query("", description="Search by filename"),
    status_filter: str = Query("all", description="all|unique|duplicate|blurred|clean|manually_blurred"),
    db: Session = Depends(get_db),
    _admin: User = Depends(require_admin),
):
    """
    Comprehensive photo registry showing ALL images with:
    - unique/duplicate status + parent image info
    - pipeline blur status, manual blur status
    - GCS folder tracking
    """
    from app.utils import get_pipeline_workspace
    workspace = get_pipeline_workspace()

    # ── Get all images from DB ──
    db_images = db.query(Image).all()
    db_by_filename = {img.filename: img for img in db_images}

    # ── Build registry entries ──
    registry = []
    for db_img in db_images:
        pipeline_blurred = db_img.compliance_status in ("blurred", "processed", "obfuscated") if db_img.compliance_status else False
        manually_blurred_flag = db_img.manually_blurred or False

        # Resolve parent image info
        parent_image_name = ""
        if db_img.parent_image_id:
            parent = db.query(Image.filename).filter(Image.id == db_img.parent_image_id).first()
            if parent:
                parent_image_name = parent[0]

        entry = {
            "filename": db_img.filename,
            "db_id": db_img.id,
            "is_unique": not db_img.is_duplicate,
            "is_duplicate": db_img.is_duplicate or False,
            "parent_image": parent_image_name,
            "pipeline_blurred": pipeline_blurred,
            "manually_blurred": manually_blurred_flag,
            "compliance_status": db_img.compliance_status or "unknown",
            "human_faces_detected": db_img.human_faces_detected or 0,
            "in_database": True,
            "is_ai_generated": db_img.is_ai_generated,
            "human_visible": db_img.human_visible,
            "original_filename": db_img.original_filename or "",
            "source_folder_id": db_img.source_folder_id or "",
            "image_drive_id": db_img.image_drive_id or "",
            "is_blurred_annotator": db_img.is_blurred_annotator or False,
            "is_restore_annotator": db_img.is_restore_annotator or False,
            "deliverable_image_path": db_img.deliverable_image_path,
            "is_manually_modified": db_img.is_manually_modified,
            "gcs_folder": db_img.gcs_folder or "input",
        }
        registry.append(entry)

    # Sort by filename
    registry.sort(key=lambda r: r["filename"])

    # ── Summary stats (before filtering) ──
    summary = {
        "total_in_drive": len(registry),
        "total_downloaded": len(registry),
        "total_unique": sum(1 for r in registry if r["is_unique"]),
        "total_duplicate": sum(1 for r in registry if r["is_duplicate"]),
        "total_pipeline_blurred": sum(1 for r in registry if r["pipeline_blurred"]),
        "total_manually_blurred": sum(1 for r in registry if r["manually_blurred"]),
        "total_in_database": len(registry),
        "total_clean": sum(1 for r in registry if not r["pipeline_blurred"] and not r["manually_blurred"] and not r["is_duplicate"]),
        "total_delivered": sum(1 for r in registry if r["deliverable_image_path"]),
    }

    # ── Apply filters ──
    if search:
        search_lower = search.lower()
        registry = [r for r in registry if search_lower in r["filename"].lower()]

    if status_filter == "unique":
        registry = [r for r in registry if r["is_unique"]]
    elif status_filter == "duplicate":
        registry = [r for r in registry if r["is_duplicate"]]
    elif status_filter == "blurred":
        registry = [r for r in registry if r["pipeline_blurred"]]
    elif status_filter == "clean":
        registry = [r for r in registry if not r["pipeline_blurred"] and not r["manually_blurred"] and not r["is_duplicate"]]
    elif status_filter == "manually_blurred":
        registry = [r for r in registry if r["manually_blurred"]]

    total = len(registry)

    # ── Paginate ──
    start = (page - 1) * per_page
    end = start + per_page
    page_data = registry[start:end]
    
    return {
        "summary": summary,
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": (total + per_page - 1) // per_page,
        "data": page_data,
    }


# ── Photo Registry Excel Export ──────────────────────────────────

@router.get("/photo-registry/export")
def export_photo_registry_excel(
    search: str = Query("", description="Search by filename"),
    status_filter: str = Query("all", description="all|unique|duplicate|blurred|clean|manually_blurred"),
    db: Session = Depends(get_db),
    _admin: User = Depends(require_admin),
):
    """
    Export the Photo Registry as an Excel (.xlsx) file.
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Reuse the same registry-building logic
    registry_response = get_photo_registry(
        page=1,
        per_page=100_000,
        search=search,
        status_filter=status_filter,
        db=db,
        _admin=_admin,
    )
    rows = registry_response["data"]
    summary = registry_response["summary"]

    # Resolve GCS bucket name
    bucket_name = os.getenv("GCS_BUCKET_NAME", "amazon-photo-pets")

    # Build GCS deliverable path
    for row in rows:
        folder_id = row.get("source_folder_id", "")
        filename = row.get("filename", "")
        gcs_folder = row.get("gcs_folder", "input")

        if folder_id and filename:
            if gcs_folder == "input":
                row["gcs_deliverable_path"] = f"gs://{bucket_name}/input/{folder_id}/{filename}"
            else:
                row["gcs_deliverable_path"] = f"gs://{bucket_name}/annotated/{folder_id}/{gcs_folder}/{filename}"
        else:
            row["gcs_deliverable_path"] = ""

    # Create workbook
    wb = Workbook()

    # ── Sheet 1: Summary ──
    ws_summary = wb.active
    ws_summary.title = "Summary"
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    summary_data = [
        ("Metric", "Value"),
        ("Total in GCS", summary.get("total_in_drive", 0)),
        ("Downloaded", summary.get("total_downloaded", 0)),
        ("Unique (After Dedup)", summary.get("total_unique", 0)),
        ("Content Duplicates", summary.get("total_duplicate", 0)),
        ("Pipeline Blurred", summary.get("total_pipeline_blurred", 0)),
        ("Manual Blur", summary.get("total_manually_blurred", 0)),
        ("Clean", summary.get("total_clean", 0)),
        ("Delivered", summary.get("total_delivered", 0)),
        ("In Database", summary.get("total_in_database", 0)),
    ]
    for r_idx, (metric, value) in enumerate(summary_data, 1):
        cell_a = ws_summary.cell(row=r_idx, column=1, value=metric)
        cell_b = ws_summary.cell(row=r_idx, column=2, value=value)
        cell_a.border = thin_border
        cell_b.border = thin_border
        if r_idx == 1:
            cell_a.font = header_font
            cell_a.fill = header_fill
            cell_b.font = header_font
            cell_b.fill = header_fill
    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 15

    # ── Sheet 2: Images ──
    ws_images = wb.create_sheet("Images")
    columns = [
        ("Filename", 30),
        ("Original Filename", 30),
        ("Image ID (Drive/GCS)", 35),
        ("Folder ID", 35),
        ("Status", 12),
        ("Compliance", 14),
        ("Pipeline Blurred", 14),
        ("Manually Blurred", 14),
        ("Is Duplicate", 12),
        ("Parent Image", 25),
        ("GCS Deliverable Path", 55),
        ("Deliverable Status", 18),
        ("In Database", 12),
    ]

    for c_idx, (col_name, width) in enumerate(columns, 1):
        cell = ws_images.cell(row=1, column=c_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border
        ws_images.column_dimensions[get_column_letter(c_idx)].width = width

    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    amber_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    red_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")

    for r_idx, row in enumerate(rows, 2):
        is_dup = row.get("is_duplicate", False)
        is_blurred = row.get("pipeline_blurred", False)

        vals = [
            row.get("filename", ""),
            row.get("original_filename", ""),
            row.get("image_drive_id", ""),
            row.get("source_folder_id", ""),
            "Duplicate" if is_dup else "Unique",
            row.get("compliance_status", ""),
            "Yes" if is_blurred else "No",
            "Yes" if row.get("manually_blurred") else "No",
            "Yes" if is_dup else "No",
            row.get("parent_image", ""),
            row.get("gcs_deliverable_path", ""),
            "Delivered" if row.get("deliverable_image_path") else "Pending",
            "Yes" if row.get("in_database") else "No",
        ]

        row_fill = amber_fill if is_dup else (red_fill if is_blurred else green_fill)
        for c_idx, val in enumerate(vals, 1):
            cell = ws_images.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_border
            cell.fill = row_fill
            cell.alignment = Alignment(vertical="center")

    ws_images.freeze_panes = "A2"
    ws_images.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename_out = f"photo_registry_{status_filter}_{len(rows)}_images.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename_out}"},
    )
